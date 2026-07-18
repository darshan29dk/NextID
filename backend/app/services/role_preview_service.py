from sqlalchemy.orm import Session
from datetime import datetime
from typing import Dict, Any, List
import json
import io
import csv

from app.models.candidate_role import CandidateRole
from app.models.candidate_role_entitlement import CandidateRoleEntitlement
from app.models.candidate_role_member import CandidateRoleMember
from app.models.role_owner_history import RoleOwnerHistory
from app.models.audit_log import AuditLog


class RolePreviewService:

    # -------------------------------------------------------------------------
    # Compute Risk Score
    # -------------------------------------------------------------------------
    @staticmethod
    def _compute_risk_score(role: CandidateRole, entitlements: list, sod_violations: list) -> int:
        """
        Computes a 0–100 composite risk score for the role based on:
          - risk_level field        (Low=10, Medium=40, High=70)
          - SoD violations          (+15 per violation, capped at 30)
          - Entitlement risk ratio  (% of high-risk entitlements × 20)
          - Classification          (Birthright=+5, Requestable=+8, Business=+10, Technical=+15)
          - Member coverage         (large roles slightly higher risk)
        """
        score = 0

        # Base from risk_level
        level_map = {"Low": 10, "Medium": 40, "High": 70}
        score += level_map.get(role.risk_level or "Low", 10)

        # SoD violations
        sod_penalty = min(len(sod_violations) * 15, 30)
        score += sod_penalty

        # Entitlement risk ratio
        if entitlements:
            high_risk_ents = [e for e in entitlements if (e.risk or "").lower() == "high"]
            ratio = len(high_risk_ents) / len(entitlements)
            score += int(ratio * 20)

        # Classification bonus — reflects how much oversight each category
        # implies: Birthright is auto-assigned (lowest), Technical carries
        # elevated/system-level access requiring the most justification.
        cls = (role.classification or "").lower()
        if cls == "birthright":
            score += 5
        elif cls == "requestable":
            score += 8
        elif cls == "business":
            score += 10
        elif cls == "technical":
            score += 15

        return min(score, 100)

    # -------------------------------------------------------------------------
    # Readiness Checks
    # -------------------------------------------------------------------------
    @staticmethod
    def _compute_readiness(
        role: CandidateRole,
        entitlements: list,
        members: list,
        primary_owner,
        sod_violations: list
    ) -> Dict:
        """
        Returns a list of validation checks with pass/fail/warning status.
        A role is 'ready' when all required checks pass.
        """
        checks = []

        # 1. Role Name
        checks.append({
            "check": "Role Name",
            "passed": bool(role.role_name and role.role_name.strip()),
            "severity": "error",
            "message": "Role must have a name" if not (role.role_name and role.role_name.strip()) else "Role name is set"
        })

        # 2. Description
        checks.append({
            "check": "Role Description",
            "passed": bool(role.role_description and role.role_description.strip()),
            "severity": "warning",
            "message": "Description is missing (recommended)" if not (role.role_description and role.role_description.strip()) else "Description is provided"
        })

        # 3. Classification
        checks.append({
            "check": "Classification",
            "passed": bool(role.classification),
            "severity": "error",
            "message": "Role must be classified before approval" if not role.classification else f"Classified as '{role.classification}'"
        })

        # 4. Primary Owner
        checks.append({
            "check": "Primary Owner",
            "passed": primary_owner is not None,
            "severity": "error",
            "message": "Primary owner must be assigned" if primary_owner is None else f"Primary owner: {primary_owner['owner_name']}"
        })

        # 5. Owner Review Date
        review_ok = False
        review_msg = "No review date set"
        if primary_owner and primary_owner.get("review_date"):
            if primary_owner.get("is_expired"):
                review_msg = "Owner review date has expired — please reassign"
                review_ok = False
            else:
                review_ok = True
                review_msg = f"Review date: {primary_owner['review_date']}"
        checks.append({
            "check": "Owner Review Date",
            "passed": review_ok,
            "severity": "warning",
            "message": review_msg
        })

        # 6. Entitlements
        core_ents = [e for e in entitlements if e.is_core]
        checks.append({
            "check": "Core Entitlements",
            "passed": len(core_ents) > 0,
            "severity": "error",
            "message": "At least one core entitlement is required" if len(core_ents) == 0 else f"{len(core_ents)} core entitlement(s) defined"
        })

        # 7. Members
        checks.append({
            "check": "Role Members",
            "passed": len(members) > 0,
            "severity": "warning",
            "message": "No members assigned to this role" if len(members) == 0 else f"{len(members)} member(s) assigned"
        })

        # 8. SoD Violations
        checks.append({
            "check": "SoD Policy",
            "passed": len(sod_violations) == 0,
            "severity": "error" if len(sod_violations) > 0 else "info",
            "message": f"{len(sod_violations)} SoD violation(s) detected — review before approval" if sod_violations else "No SoD conflicts detected"
        })

        # 9. Status progression
        valid_statuses = ["Draft", "Reviewed", "Approved"]
        checks.append({
            "check": "Status",
            "passed": role.status in valid_statuses,
            "severity": "info",
            "message": f"Current status: {role.status}"
        })

        errors = [c for c in checks if not c["passed"] and c["severity"] == "error"]
        warnings = [c for c in checks if not c["passed"] and c["severity"] == "warning"]

        return {
            "checks": checks,
            "total": len(checks),
            "passed": len([c for c in checks if c["passed"]]),
            "error_count": len(errors),
            "warning_count": len(warnings),
            "is_ready": len(errors) == 0
        }

    # -------------------------------------------------------------------------
    # Build Full Preview
    # -------------------------------------------------------------------------
    @staticmethod
    def build_preview(db: Session, role_id: int) -> Dict:
        """
        Assembles the complete Role Preview object including:
        - Role metadata
        - Computed risk score
        - Entitlements (core + non-core)
        - Members
        - Applications
        - Current owners (primary + backup)
        - Owner history
        - SoD violations
        - Audit timeline
        - Readiness checks
        """
        from sqlalchemy import or_
        from app.services.classification_service import ClassificationService

        role = db.query(CandidateRole).filter(
            CandidateRole.id == role_id,
            CandidateRole.is_deleted == False
        ).first()
        if not role:
            raise ValueError(f"Candidate role {role_id} not found")

        entitlements = db.query(CandidateRoleEntitlement).filter(
            CandidateRoleEntitlement.candidate_role_id == role_id
        ).all()

        members = db.query(CandidateRoleMember).filter(
            CandidateRoleMember.candidate_role_id == role_id
        ).all()

        # SoD
        ent_names = [e.entitlement_name for e in entitlements]
        sod_violations = ClassificationService.validate_sod_policies(ent_names)

        # Owners
        active_owners = db.query(RoleOwnerHistory).filter(
            RoleOwnerHistory.candidate_role_id == role_id,
            RoleOwnerHistory.is_active == True
        ).all()

        primary_owner = None
        backup_owner = None
        for o in active_owners:
            is_exp = o.review_date and o.review_date < datetime.utcnow()
            entry = {
                "id": o.id,
                "owner_name": o.owner_name,
                "owner_email": o.owner_email,
                "owner_user_id": o.owner_user_id,
                "owner_type": o.owner_type,
                "review_date": o.review_date.isoformat() if o.review_date else None,
                "is_expired": bool(is_exp),
                "assigned_by": o.assigned_by,
                "assigned_at": o.assigned_at.isoformat() if o.assigned_at else None,
            }
            if o.owner_type == "Primary":
                primary_owner = entry
            else:
                backup_owner = entry

        owner_history = db.query(RoleOwnerHistory).filter(
            RoleOwnerHistory.candidate_role_id == role_id
        ).order_by(RoleOwnerHistory.assigned_at.desc()).limit(20).all()

        # Risk score
        risk_score = RolePreviewService._compute_risk_score(role, entitlements, sod_violations)

        # Readiness
        readiness = RolePreviewService._compute_readiness(
            role, entitlements, members, primary_owner, sod_violations
        )

        # Audit timeline
        audit_logs = db.query(AuditLog).filter(
            or_(
                AuditLog.old_value.like(f'%"id": {role_id}%'),
                AuditLog.new_value.like(f'%"id": {role_id}%'),
                AuditLog.old_value.like(f'%"role_id": {role_id}%'),
                AuditLog.new_value.like(f'%"role_id": {role_id}%')
            )
        ).order_by(AuditLog.timestamp.desc()).limit(30).all()

        # Applications
        applications = list({e.application_name for e in entitlements if e.application_name})

        return {
            "role": {
                "id": role.id,
                "role_name": role.role_name,
                "role_description": role.role_description,
                "role_type": role.role_type,
                "risk_level": role.risk_level,
                "risk_score": risk_score,
                "classification": role.classification,
                "status": role.status,
                "confidence_score": role.confidence_score,
                "department": role.department,
                "business_unit": role.business_unit,
                "source": role.source,
                "generated_by": role.generated_by,
                "generated_on": role.generated_on.isoformat() if role.generated_on else None,
                "member_count": role.member_count,
                "user_count": role.user_count,
                "entitlement_count": role.entitlement_count,
                "application_count": role.application_count,
                "sod_violation_count": len(sod_violations),
                "created_at": role.created_at.isoformat() if role.created_at else None,
                "created_by": role.created_by
            },
            "entitlements": [
                {
                    "id": e.id,
                    "entitlement_name": e.entitlement_name,
                    "application_name": e.application_name,
                    "risk": e.risk,
                    "member_coverage_pct": e.member_coverage_pct,
                    "is_core": e.is_core
                } for e in entitlements
            ],
            "members": [
                {
                    "id": m.id,
                    "identity_id": m.identity_id,
                    "employee_id": m.employee_id,
                    "employee_name": m.employee_name,
                    "department": m.department
                } for m in members
            ],
            "applications": applications,
            "sod_violations": sod_violations,
            "owners": {
                "primary": primary_owner,
                "backup": backup_owner
            },
            "owner_history": [
                {
                    "id": h.id,
                    "owner_name": h.owner_name,
                    "owner_email": h.owner_email,
                    "owner_type": h.owner_type,
                    "review_date": h.review_date.isoformat() if h.review_date else None,
                    "is_expired": h.is_expired,
                    "is_active": h.is_active,
                    "assigned_by": h.assigned_by,
                    "assigned_at": h.assigned_at.isoformat() if h.assigned_at else None,
                    "removed_at": h.removed_at.isoformat() if h.removed_at else None,
                    "change_reason": h.change_reason
                } for h in owner_history
            ],
            "readiness": readiness,
            "generated_at": datetime.utcnow().isoformat()
        }

    # -------------------------------------------------------------------------
    # Export Preview as JSON
    # -------------------------------------------------------------------------
    @staticmethod
    def export_preview_json(db: Session, role_id: int) -> str:
        """Returns the full preview as a formatted JSON string."""
        preview = RolePreviewService.build_preview(db, role_id)
        return json.dumps(preview, indent=2, default=str)

    # -------------------------------------------------------------------------
    # Export Preview as CSV (multi-sheet represented as sections)
    # -------------------------------------------------------------------------
    @staticmethod
    def export_preview_csv(db: Session, role_id: int) -> str:
        """Returns a UTF-8 CSV string with multiple sections for the preview."""
        preview = RolePreviewService.build_preview(db, role_id)
        output = io.StringIO()
        writer = csv.writer(output)

        # --- Role Metadata ---
        writer.writerow(["=== ROLE METADATA ==="])
        role = preview["role"]
        for key, val in role.items():
            writer.writerow([key, val])
        writer.writerow([])

        # --- Owners ---
        writer.writerow(["=== ROLE OWNERS ==="])
        writer.writerow(["Type", "Name", "Email", "Review Date", "Expired", "Assigned By", "Assigned At"])
        for owner_type, owner_data in [("Primary", preview["owners"]["primary"]), ("Backup", preview["owners"]["backup"])]:
            if owner_data:
                writer.writerow([
                    owner_data["owner_type"],
                    owner_data["owner_name"],
                    owner_data.get("owner_email") or "",
                    owner_data.get("review_date") or "",
                    "Yes" if owner_data.get("is_expired") else "No",
                    owner_data.get("assigned_by") or "",
                    owner_data.get("assigned_at") or ""
                ])
            else:
                writer.writerow([owner_type, "Not Assigned", "", "", "", "", ""])
        writer.writerow([])

        # --- Entitlements ---
        writer.writerow(["=== ENTITLEMENTS ==="])
        writer.writerow(["Entitlement Name", "Application", "Risk", "Coverage %", "Core"])
        for e in preview["entitlements"]:
            writer.writerow([
                e["entitlement_name"], e["application_name"] or "", e["risk"] or "",
                e["member_coverage_pct"] or "", "Yes" if e["is_core"] else "No"
            ])
        writer.writerow([])

        # --- Members ---
        writer.writerow(["=== MEMBERS ==="])
        writer.writerow(["Employee ID", "Name", "Department"])
        for m in preview["members"]:
            writer.writerow([m["employee_id"] or "", m["employee_name"] or "", m["department"] or ""])
        writer.writerow([])

        # --- SoD Violations ---
        writer.writerow(["=== SOD VIOLATIONS ==="])
        if preview["sod_violations"]:
            writer.writerow(["Entitlement 1", "Entitlement 2", "Description"])
            for v in preview["sod_violations"]:
                writer.writerow([v["entitlement_1"], v["entitlement_2"], v["description"]])
        else:
            writer.writerow(["No SoD violations detected"])
        writer.writerow([])

        # --- Readiness Checks ---
        writer.writerow(["=== READINESS CHECKS ==="])
        writer.writerow(["Check", "Passed", "Severity", "Message"])
        for c in preview["readiness"]["checks"]:
            writer.writerow([c["check"], "Yes" if c["passed"] else "No", c["severity"], c["message"]])

        output.seek(0)
        return output.getvalue()

    # -------------------------------------------------------------------------
    # Export Preview as Excel
    # -------------------------------------------------------------------------
    @staticmethod
    def export_preview_excel(db: Session, role_id: int) -> bytes:
        """Returns an Excel workbook bytes for the role preview with multiple sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        preview = RolePreviewService.build_preview(db, role_id)
        wb = Workbook()

        # Helper: style header row
        HEADER_FILL = PatternFill("solid", fgColor="1A3557")
        HEADER_FONT = Font(color="FFFFFF", bold=True)

        def style_headers(ws, row_idx=1):
            for cell in ws[row_idx]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # --- Sheet 1: Role Overview ---
        ws1 = wb.active
        ws1.title = "Role Overview"
        role = preview["role"]
        ws1.append(["Field", "Value"])
        style_headers(ws1)
        for key, val in role.items():
            ws1.append([key.replace("_", " ").title(), str(val) if val is not None else ""])

        # Readiness summary
        ws1.append([])
        ws1.append(["=== Readiness Summary ===", ""])
        r = preview["readiness"]
        ws1.append(["Total Checks", r["total"]])
        ws1.append(["Passed", r["passed"]])
        ws1.append(["Errors", r["error_count"]])
        ws1.append(["Warnings", r["warning_count"]])
        ws1.append(["Ready for Approval", "Yes" if r["is_ready"] else "No"])
        auto_width(ws1)

        # --- Sheet 2: Owners ---
        ws2 = wb.create_sheet("Owners")
        ws2.append(["Type", "Name", "Email", "Review Date", "Expired", "Assigned By", "Assigned At"])
        style_headers(ws2)
        for owner_key in ["primary", "backup"]:
            o = preview["owners"][owner_key]
            if o:
                ws2.append([
                    o["owner_type"], o["owner_name"], o.get("owner_email") or "",
                    o.get("review_date") or "", "Yes" if o.get("is_expired") else "No",
                    o.get("assigned_by") or "", o.get("assigned_at") or ""
                ])
            else:
                ws2.append([owner_key.capitalize(), "Not Assigned", "", "", "", "", ""])

        # Owner History sub-section
        ws2.append([])
        ws2.append(["=== Owner History ===", "", "", "", "", "", ""])
        ws2.append(["Type", "Name", "Email", "Review Date", "Active", "Assigned By", "Removed At"])
        for h in preview["owner_history"]:
            ws2.append([
                h["owner_type"], h["owner_name"], h.get("owner_email") or "",
                h.get("review_date") or "", "Yes" if h["is_active"] else "No",
                h.get("assigned_by") or "", h.get("removed_at") or ""
            ])
        auto_width(ws2)

        # --- Sheet 3: Entitlements ---
        ws3 = wb.create_sheet("Entitlements")
        ws3.append(["Entitlement Name", "Application", "Risk Level", "Coverage %", "Core"])
        style_headers(ws3)
        for e in preview["entitlements"]:
            ws3.append([
                e["entitlement_name"], e["application_name"] or "", e["risk"] or "",
                e["member_coverage_pct"] or 0, "Yes" if e["is_core"] else "No"
            ])
        auto_width(ws3)

        # --- Sheet 4: Members ---
        ws4 = wb.create_sheet("Members")
        ws4.append(["Employee ID", "Name", "Department"])
        style_headers(ws4)
        for m in preview["members"]:
            ws4.append([m["employee_id"] or "", m["employee_name"] or "", m["department"] or ""])
        auto_width(ws4)

        # --- Sheet 5: SoD Violations ---
        ws5 = wb.create_sheet("SoD Violations")
        if preview["sod_violations"]:
            ws5.append(["Entitlement 1", "Entitlement 2", "Description"])
            style_headers(ws5)
            for v in preview["sod_violations"]:
                ws5.append([v["entitlement_1"], v["entitlement_2"], v["description"]])
        else:
            ws5.append(["No SoD violations detected"])
        auto_width(ws5)

        # --- Sheet 6: Readiness Checks ---
        ws6 = wb.create_sheet("Readiness Checks")
        ws6.append(["Check", "Passed", "Severity", "Message"])
        style_headers(ws6)
        for c in preview["readiness"]["checks"]:
            ws6.append([c["check"], "Yes" if c["passed"] else "No", c["severity"], c["message"]])
        auto_width(ws6)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
