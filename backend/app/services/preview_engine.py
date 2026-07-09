import os
import json
from datetime import datetime
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional

from app.models.connector import Connector
from app.models.connector_field_mapping import ConnectorFieldMapping
from app.models.transformation_rule import TransformationRule
from app.models.validation_rule import ValidationRule
from app.models.import_preview import ImportPreview
from app.services.transformation_engine import TransformationEngine
from app.services.validation_engine import ValidationEngine

class PreviewEngine:
    @staticmethod
    def generate_preview(db: Session, connector_id: int, table_name: Optional[str] = None) -> Dict[str, Any]:
        # 1. Fetch connector
        connector = db.query(Connector).filter(Connector.id == connector_id, Connector.is_deleted == False).first()
        if not connector:
            raise Exception("Connector not found")

        # 2. Clear existing preview cache for this connector
        db.query(ImportPreview).filter(ImportPreview.connector_id == connector_id).delete()
        db.commit()

        # 3. Read raw records (up to 100 for dry-run preview)
        raw_rows = PreviewEngine._read_raw_records(connector, table_name)
        if not raw_rows:
            return {"message": "No records found to preview", "count": 0}

        # 4. Load mappings, transformations, and validation rules
        mappings = db.query(ConnectorFieldMapping).filter(ConnectorFieldMapping.connector_id == connector_id).all()
        transformations = db.query(TransformationRule).filter(
            TransformationRule.connector_id == connector_id,
            TransformationRule.enabled == True,
            TransformationRule.is_deleted == False
        ).order_by(TransformationRule.execution_order.asc()).all()

        validations = db.query(ValidationRule).filter(
            ValidationRule.connector_id == connector_id,
            ValidationRule.enabled == True,
            ValidationRule.is_deleted == False
        ).order_by(ValidationRule.execution_order.asc()).all()

        # Build mapping lookups
        mapping_by_id = {m.id: m for m in mappings}
        # Find which target attribute names correspond to which source fields
        source_to_target = {m.source_field: m.target_attribute_name for m in mappings}

        # Seen values map for Unique validation rule trackers
        seen_values_by_rule = {rule.id: set() for rule in validations if rule.validation_type.lower() == "unique"}

        # 5. Process records
        preview_records = []
        for index, raw_row in enumerate(raw_rows):
            record_number = index + 1
            
            # Map raw fields to initial target attributes
            original_mapped = {}
            for source_f, val in raw_row.items():
                target_attr = source_to_target.get(source_f)
                if target_attr:
                    original_mapped[target_attr] = str(val) if val is not None else ""

            # Fill in missing mapped fields as empty
            for m in mappings:
                if m.target_attribute_name not in original_mapped:
                    original_mapped[m.target_attribute_name] = ""

            transformed = dict(original_mapped)

            # Apply transformations
            for trans_rule in transformations:
                if trans_rule.mapping_id in mapping_by_id:
                    mapping = mapping_by_id[trans_rule.mapping_id]
                    target_attr = mapping.target_attribute_name
                    current_val = transformed.get(target_attr, "")
                    
                    # Apply transformation logic passing the raw row data for any concatenations / expressions
                    transformed_val = TransformationEngine.transform_value(
                        value=current_val,
                        rule_type=trans_rule.transformation_type,
                        expression=trans_rule.expression,
                        parameters_str=trans_rule.parameters,
                        row_data=raw_row
                    )
                    transformed[target_attr] = transformed_val

            # Apply validations
            record_status = "Valid"
            record_errors = []
            record_warnings = []
            field_validations = {}  # Field-level results: {field_name: {status, message}}

            for val_rule in validations:
                if val_rule.mapping_id in mapping_by_id:
                    mapping = mapping_by_id[val_rule.mapping_id]
                    target_attr = mapping.target_attribute_name
                    transformed_val = transformed.get(target_attr, "")

                    # Unique rule requires passing the seen set tracker
                    seen_set = seen_values_by_rule.get(val_rule.id)
                    
                    val_res = ValidationEngine.validate_value(
                        value=transformed_val,
                        validation_type=val_rule.validation_type,
                        parameters_str=val_rule.parameters,
                        error_message=val_rule.error_message,
                        seen_values=seen_set,
                        severity=val_rule.severity
                    )

                    if not val_res["valid"]:
                        status = val_res["status"]  # "Error", "Warning", "Info"
                        msg = val_res["message"]
                        
                        # Populate field-level structure
                        if target_attr not in field_validations:
                            field_validations[target_attr] = []
                        field_validations[target_attr].append({
                            "status": status,
                            "message": msg,
                            "rule_name": val_rule.rule_name
                        })

                        # Update record level summaries
                        if status == "Error":
                            record_errors.append(f"{target_attr}: {msg}")
                            record_status = "Error"
                        elif status == "Warning":
                            record_warnings.append(f"{target_attr}: {msg}")
                            # Error takes precedence over Warning status
                            if record_status != "Error":
                                record_status = "Warning"

            # Create Database preview record
            preview_rec = ImportPreview(
                connector_id=connector_id,
                record_number=record_number,
                source_data=json.dumps(original_mapped),
                transformed_data=json.dumps(transformed),
                validation_result=json.dumps(field_validations),
                status=record_status,
                errors=json.dumps(record_errors),
                warnings=json.dumps(record_warnings),
                previewed_at=datetime.utcnow()
            )
            db.add(preview_rec)
            preview_records.append(preview_rec)

        db.commit()
        return {"message": "Preview generated successfully", "count": len(preview_records)}

    @staticmethod
    def _read_raw_records(connector: Connector, table_name: Optional[str] = None, limit: Optional[int] = 100) -> List[Dict[str, Any]]:
        raw_rows = []

        if connector.connector_type in ["CSV", "Excel"]:
            if not connector.file_path:
                raise Exception("No file has been uploaded to this connector")
            
            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                connector.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"Connector source file not found: {connector.file_path}")

            if connector.connector_type == "CSV":
                import csv as csv_module
                encoding = connector.csv_encoding or "UTF-8"
                delimiter = connector.csv_delimiter or ","
                if delimiter == "	":
                    delimiter = "\t"
                
                with open(full_path, "r", encoding=encoding) as f:
                    # Read lines dynamically
                    reader = csv_module.DictReader(f, delimiter=delimiter)
                    for row in reader:
                        # Clean key spacing
                        stripped_row = {str(k).strip(): v for k, v in row.items()}
                        raw_rows.append(stripped_row)
                        if limit is not None and len(raw_rows) >= limit:  # Dynamic limit
                            break

            elif connector.connector_type == "Excel":
                import openpyxl
                wb = openpyxl.load_workbook(full_path, data_only=True, read_only=True)
                sheet = wb[connector.excel_sheet_name] if connector.excel_sheet_name and connector.excel_sheet_name in wb.sheetnames else wb.active
                
                rows_iter = sheet.iter_rows(values_only=True)
                headers = next(rows_iter, None)
                if not headers:
                    raise Exception("Excel worksheet appears empty")
                
                headers = [str(h).strip() for h in headers if h is not None]
                
                for row in rows_iter:
                    if all(val is None for val in row):
                        continue
                    row_dict = {}
                    for idx, h in enumerate(headers):
                        if idx < len(row):
                            row_dict[h] = row[idx]
                    raw_rows.append(row_dict)
                    if limit is not None and len(raw_rows) >= limit:
                        break

        elif connector.connector_type == "Database":
            if connector.database_type != "MySQL":
                raise Exception("Only MySQL database connector is supported for schema previews currently")
            if not table_name:
                raise Exception("A valid table name is required to run database preview")
            
            import pymysql
            from app.utils.crypto import decrypt_password
            
            decrypted_pw = decrypt_password(connector.password) if connector.password else ""
            conn = pymysql.connect(
                host=connector.host,
                port=connector.port or 3306,
                user=connector.username,
                password=decrypted_pw,
                database=connector.database_name,
                connect_timeout=connector.connection_timeout or 30
            )
            try:
                with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                    # Basic SQL validation escaping backticks safely
                    limit_sql = f" LIMIT {limit}" if limit is not None else ""
                    cursor.execute(f"SELECT * FROM `{table_name}`{limit_sql}")
                    raw_rows = cursor.fetchall()
            finally:
                conn.close()

        elif connector.connector_type == "API Gateway":
            import requests as requests_lib
            from app.utils.crypto import decrypt_password
            
            url = connector.host
            if not url:
                raise Exception("API Gateway URL is required.")
            
            headers = {}
            if connector.file_path:
                try:
                    headers = json.loads(connector.file_path)
                except Exception:
                    pass

            auth = None
            if connector.auth_type == "Basic":
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                auth = (connector.username, decrypted_pw)
            elif connector.auth_type == "API Key" and connector.username and connector.password:
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                headers[connector.username] = decrypted_pw

            timeout = connector.connection_timeout or 30
            resp = requests_lib.get(url, headers=headers, auth=auth, timeout=timeout)
            if resp.status_code >= 400:
                raise Exception(f"HTTP Error {resp.status_code}: {resp.text[:100]}")
            
            res_json = resp.json()
            json_key = connector.database_name
            target_list = res_json
            if json_key:
                if json_key in res_json:
                    target_list = res_json[json_key]
            
            if not isinstance(target_list, list):
                if isinstance(target_list, dict):
                    target_list = [target_list]
                else:
                    target_list = []
                    
            for item in target_list:
                if isinstance(item, dict):
                    raw_rows.append({str(k).strip(): v for k, v in item.items()})
                if limit is not None and len(raw_rows) >= limit:
                    break

        return raw_rows
