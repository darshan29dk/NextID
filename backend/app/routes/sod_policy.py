import json
import csv
import io
import re
from fastapi import APIRouter, HTTPException, Depends, Header, status, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from typing import List, Optional
from datetime import datetime
import openpyxl

from app.database import get_db
from app.utils.permissions import require_permission
from app.models.sod_policy import SodPolicy, SodPolicyRule, SodPolicyAudit
from app.models.application import Application
from app.models.application_entitlement import ApplicationEntitlement
from app.models.identity import Identity
from app.models.application_account import ApplicationAccount
from app.models.application_account_entitlement import ApplicationAccountEntitlement
from app.models.audit_log import AuditLog
from app.models.sod_violation import SodViolation, SodScanHistory, SodViolationAudit, SodViolationComment, SodViolationAttachment
from app.schemas.sod_policy import (
    SodPolicyCreate,
    SodPolicyUpdate,
    SodPolicyResponse,
    SodPolicyAuditResponse,
    SodPolicyListResponse
)

router = APIRouter()

def write_sod_audit(db: Session, policy_id: str, action: str, performed_by: str, old_val: dict = None, new_val: dict = None):
    """Writes an audit entry to both the specialized sod_policy_audit and global platform audit_logs."""
    old_str = json.dumps(old_val) if old_val else None
    new_str = json.dumps(new_val) if new_val else None
    
    # 1. Sod-specific audit
    sod_audit = SodPolicyAudit(
        policy_id=policy_id,
        action=action,
        performed_by=performed_by,
        old_value=old_str,
        new_value=new_str,
        timestamp=datetime.utcnow()
    )
    db.add(sod_audit)
    
    # 2. Global platform audit
    global_audit = AuditLog(
        module="Governance",
        action=f"SoD Policy {action}",
        performed_by=performed_by,
        old_value=old_str,
        new_value=new_str,
        timestamp=datetime.utcnow()
    )
    db.add(global_audit)
    db.commit()

def check_duplicate_rules(db: Session, rules_data: list, exclude_policy_id: str = None) -> Optional[str]:
    """Checks if there is another policy that has the exact same set of entitlement rules."""
    # Convert incoming rules to a canonical sorted tuple representation
    incoming_rules = sorted([
        (r.application_name, r.entitlement_one, r.entitlement_two, r.condition_type)
        for r in rules_data
    ])
    
    # Query all active policies
    policies = db.query(SodPolicy).filter(SodPolicy.status != "DEPRECATED").all()
    for p in policies:
        if exclude_policy_id and p.id == exclude_policy_id:
            continue
        p_rules = sorted([
            (r.application_name, r.entitlement_one, r.entitlement_two, r.condition_type)
            for r in p.rules
        ])
        if incoming_rules == p_rules:
            return p.policy_code
            
    return None

def get_next_policy_code(db: Session) -> str:
    """Generates the next logical policy code like SOD-001, SOD-002, etc."""
    count = db.query(SodPolicy).count()
    return f"SOD-{str(count + 1).zfill(3)}"

@router.get("/governance/sod-policies/lookup/applications", response_model=List[str])
def get_lookup_applications(db: Session = Depends(get_db)):
    """Returns a list of all configured application names."""
    apps = db.query(Application.application_name).filter(Application.is_deleted == False).all()
    return [a[0] for a in apps]

@router.get("/governance/sod-policies/lookup/entitlements", response_model=List[str])
def get_lookup_entitlements(application_name: str, db: Session = Depends(get_db)):
    """Returns a list of entitlement names for a specific application."""
    app = db.query(Application).filter(Application.application_name == application_name, Application.is_deleted == False).first()
    if not app:
        return []
    ents = db.query(ApplicationEntitlement.entitlement_name).filter(
        ApplicationEntitlement.application_id == app.id,
        ApplicationEntitlement.is_deleted == False
    ).all()
    return [e[0] for e in ents]

@router.post("/governance/sod-policies", response_model=SodPolicyResponse, dependencies=[Depends(require_permission("SoD Policies", "create"))])
def create_sod_policy(
    payload: SodPolicyCreate,
    x_user_name: str = Header(default="System"),
    db: Session = Depends(get_db)
):
    # Validate rules count
    if not payload.rules:
        raise HTTPException(status_code=400, detail="SoD policy must contain at least one rule row.")
        
    # Check duplicate rule definition
    duplicate_code = check_duplicate_rules(db, payload.rules)
    if duplicate_code:
        raise HTTPException(
            status_code=400, 
            detail=f"Duplicate policy rule detected. An identical set of rules already exists in policy '{duplicate_code}'."
        )
        
    # Generate next code
    code = get_next_policy_code(db)
    
    # Create policy
    policy = SodPolicy(
        policy_code=code,
        policy_name=payload.policy_name,
        description=payload.description,
        risk_level=payload.risk_level,
        policy_type=payload.policy_type,
        status=payload.status,
        business_owner=payload.business_owner,
        approver=payload.approver,
        created_by=x_user_name,
        version=1
    )
    db.add(policy)
    db.flush()
    
    # Add rules
    for r in payload.rules:
        rule = SodPolicyRule(
            policy_id=policy.id,
            application_name=r.application_name,
            entitlement_one=r.entitlement_one,
            entitlement_two=r.entitlement_two,
            condition_type=r.condition_type
        )
        db.add(rule)
        
    db.commit()
    db.refresh(policy)
    
    # Audit log
    new_val = {
        "policy_code": policy.policy_code,
        "policy_name": policy.policy_name,
        "rules": [{"app": r.application_name, "ent1": r.entitlement_one, "ent2": r.entitlement_two, "op": r.condition_type} for r in policy.rules]
    }
    write_sod_audit(db, policy.id, "Create", x_user_name, new_val=new_val)
    
    return policy

@router.get("/governance/sod-policies", response_model=SodPolicyListResponse)
def get_sod_policies(
    search: Optional[str] = None,
    risk_level: Optional[str] = None,
    status: Optional[str] = None,
    policy_type: Optional[str] = None,
    page: int = 1,
    limit: int = 10,
    db: Session = Depends(get_db)
):
    query = db.query(SodPolicy)
    
    # Apply search filter
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                SodPolicy.policy_name.like(search_term),
                SodPolicy.policy_code.like(search_term),
                SodPolicy.business_owner.like(search_term)
            )
        )
        
    # Apply advanced filters
    if risk_level:
        query = query.filter(SodPolicy.risk_level == risk_level)
    if status:
        query = query.filter(SodPolicy.status == status)
    if policy_type:
        query = query.filter(SodPolicy.policy_type == policy_type)
        
    total = query.count()
    
    # Paginate
    policies = query.order_by(SodPolicy.created_date.desc()).offset((page - 1) * limit).limit(limit).all()
    
    # Aggregate KPIs for dashboard
    kpis = {
        "total": db.query(SodPolicy).count(),
        "active": db.query(SodPolicy).filter(SodPolicy.status == "ACTIVE").count(),
        "inactive": db.query(SodPolicy).filter(SodPolicy.status == "INACTIVE").count(),
        "critical": db.query(SodPolicy).filter(SodPolicy.risk_level == "CRITICAL").count(),
        "high": db.query(SodPolicy).filter(SodPolicy.risk_level == "HIGH").count(),
        "draft": db.query(SodPolicy).filter(SodPolicy.status == "DRAFT").count()
    }
    
    return {
        "policies": policies,
        "total": total,
        "page": page,
        "limit": limit,
        "kpis": kpis
    }

@router.get("/governance/sod-policies/{id}", response_model=SodPolicyResponse)
def get_sod_policy(id: str, db: Session = Depends(get_db)):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
    return policy

@router.get("/governance/sod-policies/{id}/audit", response_model=List[SodPolicyAuditResponse])
def get_sod_policy_audit(id: str, db: Session = Depends(get_db)):
    audits = db.query(SodPolicyAudit).filter(SodPolicyAudit.policy_id == id).order_by(SodPolicyAudit.timestamp.desc()).all()
    return audits

@router.put("/governance/sod-policies/{id}", response_model=SodPolicyResponse, dependencies=[Depends(require_permission("SoD Policies", "edit"))])
def update_sod_policy(
    id: str,
    payload: SodPolicyUpdate,
    x_user_name: str = Header(default="System"),
    db: Session = Depends(get_db)
):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
        
    # Validate rules count
    if not payload.rules:
        raise HTTPException(status_code=400, detail="SoD policy must contain at least one rule row.")
        
    # Check duplicate rule definition
    duplicate_code = check_duplicate_rules(db, payload.rules, exclude_policy_id=id)
    if duplicate_code:
        raise HTTPException(
            status_code=400, 
            detail=f"Duplicate policy rule detected. An identical set of rules already exists in policy '{duplicate_code}'."
        )
        
    # Capture old value for audit logging
    old_val = {
        "policy_name": policy.policy_name,
        "description": policy.description,
        "risk_level": policy.risk_level,
        "policy_type": policy.policy_type,
        "status": policy.status,
        "rules": [{"app": r.application_name, "ent1": r.entitlement_one, "ent2": r.entitlement_two, "op": r.condition_type} for r in policy.rules]
    }
    
    # Update policy attributes
    policy.policy_name = payload.policy_name
    policy.description = payload.description
    policy.risk_level = payload.risk_level
    policy.policy_type = payload.policy_type
    policy.status = payload.status
    policy.business_owner = payload.business_owner
    policy.approver = payload.approver
    policy.updated_by = x_user_name
    policy.updated_date = datetime.utcnow()
    policy.version += 1
    
    # Recreate rules list
    db.query(SodPolicyRule).filter(SodPolicyRule.policy_id == id).delete()
    for r in payload.rules:
        rule = SodPolicyRule(
            policy_id=policy.id,
            application_name=r.application_name,
            entitlement_one=r.entitlement_one,
            entitlement_two=r.entitlement_two,
            condition_type=r.condition_type
        )
        db.add(rule)
        
    db.commit()
    db.refresh(policy)
    
    # Audit log
    new_val = {
        "policy_name": policy.policy_name,
        "description": policy.description,
        "risk_level": policy.risk_level,
        "policy_type": policy.policy_type,
        "status": policy.status,
        "rules": [{"app": r.application_name, "ent1": r.entitlement_one, "ent2": r.entitlement_two, "op": r.condition_type} for r in policy.rules]
    }
    write_sod_audit(db, policy.id, "Update", x_user_name, old_val=old_val, new_val=new_val)
    
    return policy

@router.delete("/governance/sod-policies/{id}", response_model=dict, dependencies=[Depends(require_permission("SoD Policies", "delete"))])
def delete_sod_policy(id: str, x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
        
    # Capture old state for audit logging
    old_val = {"policy_code": policy.policy_code, "policy_name": policy.policy_name}

    # SodViolation.policy_id is a NOT NULL FK with no ON DELETE CASCADE, so
    # deleting a policy that already has violations against it fails with an
    # IntegrityError (surfaced to the UI as a generic "Failed to delete
    # policy"). Clean up violations - and their comments/attachments/audit
    # rows - before deleting the policy itself.
    violation_ids = [
        v_id for (v_id,) in db.query(SodViolation.id).filter(SodViolation.policy_id == id).all()
    ]
    if violation_ids:
        db.query(SodViolationAudit).filter(SodViolationAudit.violation_id.in_(violation_ids)).delete(synchronize_session=False)
        db.query(SodViolationComment).filter(SodViolationComment.violation_id.in_(violation_ids)).delete(synchronize_session=False)
        db.query(SodViolationAttachment).filter(SodViolationAttachment.violation_id.in_(violation_ids)).delete(synchronize_session=False)
        db.query(SodViolation).filter(SodViolation.policy_id == id).delete(synchronize_session=False)

    db.delete(policy)
    db.commit()
    
    write_sod_audit(db, id, "Delete", x_user_name, old_val=old_val)
    
    return {"message": "SoD Policy deleted successfully"}

@router.post("/governance/sod-policies/{id}/clone", response_model=SodPolicyResponse, dependencies=[Depends(require_permission("SoD Policies", "create"))])
def clone_sod_policy(id: str, x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
        
    # Generate new code
    code = get_next_policy_code(db)
    
    # Create duplicate
    cloned = SodPolicy(
        policy_code=code,
        policy_name=f"Copy of {policy.policy_name}",
        description=policy.description,
        risk_level=policy.risk_level,
        policy_type=policy.policy_type,
        status="DRAFT",
        business_owner=policy.business_owner,
        approver=policy.approver,
        created_by=x_user_name,
        version=1
    )
    db.add(cloned)
    db.flush()
    
    # Clone rules
    for r in policy.rules:
        rule = SodPolicyRule(
            policy_id=cloned.id,
            application_name=r.application_name,
            entitlement_one=r.entitlement_one,
            entitlement_two=r.entitlement_two,
            condition_type=r.condition_type
        )
        db.add(rule)
        
    db.commit()
    db.refresh(cloned)
    
    new_val = {"policy_code": cloned.policy_code, "policy_name": cloned.policy_name}
    write_sod_audit(db, cloned.id, "Clone", x_user_name, new_val=new_val)
    
    return cloned

@router.patch("/governance/sod-policies/{id}/activate", response_model=SodPolicyResponse, dependencies=[Depends(require_permission("SoD Policies", "edit"))])
def activate_sod_policy(id: str, x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
        
    old_status = policy.status
    policy.status = "ACTIVE"
    policy.updated_by = x_user_name
    policy.updated_date = datetime.utcnow()
    db.commit()
    
    write_sod_audit(
        db, 
        policy.id, 
        "Activate", 
        x_user_name, 
        old_val={"status": old_status}, 
        new_val={"status": "ACTIVE"}
    )
    
    return policy

@router.patch("/governance/sod-policies/{id}/deactivate", response_model=SodPolicyResponse, dependencies=[Depends(require_permission("SoD Policies", "edit"))])
def deactivate_sod_policy(id: str, x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
        
    old_status = policy.status
    policy.status = "INACTIVE"
    policy.updated_by = x_user_name
    policy.updated_date = datetime.utcnow()
    db.commit()
    
    write_sod_audit(
        db, 
        policy.id, 
        "Deactivate", 
        x_user_name, 
        old_val={"status": old_status}, 
        new_val={"status": "INACTIVE"}
    )
    
    return policy

# ── Bulk Actions ──
@router.post("/governance/sod-policies/bulk-delete", dependencies=[Depends(require_permission("SoD Policies", "delete"))])
def bulk_delete_sod_policies(ids: List[str], x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policies = db.query(SodPolicy).filter(SodPolicy.id.in_(ids)).all()
    count = len(policies)
    for p in policies:
        db.delete(p)
        write_sod_audit(db, p.id, "Delete (Bulk)", x_user_name, old_val={"policy_code": p.policy_code})
    db.commit()
    return {"message": f"Successfully deleted {count} policies."}

@router.post("/governance/sod-policies/bulk-activate", dependencies=[Depends(require_permission("SoD Policies", "edit"))])
def bulk_activate_sod_policies(ids: List[str], x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policies = db.query(SodPolicy).filter(SodPolicy.id.in_(ids)).all()
    count = len(policies)
    for p in policies:
        old_status = p.status
        p.status = "ACTIVE"
        p.updated_by = x_user_name
        p.updated_date = datetime.utcnow()
        write_sod_audit(db, p.id, "Activate (Bulk)", x_user_name, old_val={"status": old_status}, new_val={"status": "ACTIVE"})
    db.commit()
    return {"message": f"Successfully activated {count} policies."}

@router.post("/governance/sod-policies/bulk-deactivate", dependencies=[Depends(require_permission("SoD Policies", "edit"))])
def bulk_deactivate_sod_policies(ids: List[str], x_user_name: str = Header(default="System"), db: Session = Depends(get_db)):
    policies = db.query(SodPolicy).filter(SodPolicy.id.in_(ids)).all()
    count = len(policies)
    for p in policies:
        old_status = p.status
        p.status = "INACTIVE"
        p.updated_by = x_user_name
        p.updated_date = datetime.utcnow()
        write_sod_audit(db, p.id, "Deactivate (Bulk)", x_user_name, old_val={"status": old_status}, new_val={"status": "INACTIVE"})
    db.commit()
    return {"message": f"Successfully deactivated {count} policies."}

# ── Simulation & Impact Analysis Engine ──
def compute_violations(db: Session, rules: List[SodPolicyRule]) -> List[dict]:
    """Simulates policies and calculates violating identities from active entitlements."""
    if not rules:
        return []
        
    # Group rules by Application
    # Rule holds entitlement_one and entitlement_two with condition_type (AND, OR, NOT)
    violations = []
    
    # Query all active identities
    identities = db.query(Identity).filter(Identity.is_deleted == False).all()
    for ident in identities:
        # Load all entitlements held by this identity
        # identity -> accounts -> entitlements
        accounts = db.query(ApplicationAccount).filter(
            ApplicationAccount.identity_id == ident.id,
            ApplicationAccount.is_deleted == False
        ).all()
        
        # Build dictionary of application name -> set of entitlement names
        held_ents = {}
        for acc in accounts:
            app = db.query(Application).filter(Application.id == acc.application_id).first()
            if not app:
                continue
            app_name = app.application_name
            if app_name not in held_ents:
                held_ents[app_name] = set()
                
            links = db.query(ApplicationAccountEntitlement).filter(
                ApplicationAccountEntitlement.account_id == acc.id
            ).all()
            for lnk in links:
                held_ents[app_name].add(lnk.entitlement_name_raw.strip().lower())
                
        # Evaluate rules
        is_violator = False
        violating_details = []
        
        for rule in rules:
            app = rule.application_name
            ent1 = rule.entitlement_one.strip().lower()
            ent2 = rule.entitlement_two.strip().lower()
            op = rule.condition_type
            
            # Check if this application has entitlements for the user
            user_app_ents = held_ents.get(app, set())
            
            has_ent1 = ent1 in user_app_ents
            has_ent2 = ent2 in user_app_ents
            
            violated_rule = False
            if op == "AND":
                violated_rule = has_ent1 and has_ent2
            elif op == "OR":
                violated_rule = has_ent1 or has_ent2
            elif op == "NOT":
                # User has entitlement_one but NOT entitlement_two
                violated_rule = has_ent1 and not has_ent2
                
            if violated_rule:
                is_violator = True
                violating_details.append({
                    "application": app,
                    "entitlement_one": rule.entitlement_one,
                    "entitlement_two": rule.entitlement_two,
                    "operator": op
                })
                
        if is_violator:
            violations.append({
                "id": ident.id,
                "employee_id": ident.employee_id,
                "name": ident.display_name or f"{ident.first_name or ''} {ident.last_name or ''}".strip(),
                "email": ident.email,
                "department": ident.department,
                "violations": violating_details
            })
            
    return violations

@router.post("/governance/sod-policies/simulate")
def simulate_sod_policy(payload: SodPolicyCreate, db: Session = Depends(get_db)):
    """Simulates a brand new unsaved policy and returns projected violations list."""
    rules = [
        SodPolicyRule(
            application_name=r.application_name,
            entitlement_one=r.entitlement_one,
            entitlement_two=r.entitlement_two,
            condition_type=r.condition_type
        )
        for r in payload.rules
    ]
    violators = compute_violations(db, rules)
    return {
        "violators_count": len(violators),
        "violators": violators
    }

@router.get("/governance/sod-policies/{id}/simulate")
def simulate_existing_policy(id: str, db: Session = Depends(get_db)):
    """Runs simulation on an existing saved policy."""
    policy = db.query(SodPolicy).filter(SodPolicy.id == id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="SoD policy not found.")
    violators = compute_violations(db, policy.rules)
    return {
        "violators_count": len(violators),
        "violators": violators
    }

# ── Bulk Imports (JSON, CSV, Excel) ──
def parse_rule_expression_string(rule_str: str, default_app: str = "Global"):
    """
    Parses strings like:
    - 'Salesforce: Create Invoices AND Approve Invoices'
    - 'Create Invoices AND Approve Invoices'
    - 'Create Invoices vs Approve Invoices'
    - 'Create Invoices & Approve Invoices'
    - 'Create Invoices, Approve Invoices'
    """
    if not isinstance(rule_str, str):
        return default_app, "Entitlement A", "Entitlement B", "AND"

    rule_str = rule_str.strip()
    app_name = default_app

    if ":" in rule_str and not rule_str.startswith("http"):
        parts = rule_str.split(":", 1)
        if len(parts[0].strip()) > 0 and len(parts[1].strip()) > 0:
            app_name = parts[0].strip()
            rule_str = parts[1].strip()

    op = "AND"
    ent1 = ""
    ent2 = ""

    if " AND " in rule_str.upper():
        parts = re.split(r'\s+AND\s+', rule_str, flags=re.IGNORECASE)
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "AND"
    elif " OR " in rule_str.upper():
        parts = re.split(r'\s+OR\s+', rule_str, flags=re.IGNORECASE)
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "OR"
    elif " NOT " in rule_str.upper():
        parts = re.split(r'\s+NOT\s+', rule_str, flags=re.IGNORECASE)
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "NOT"
    elif " VS " in rule_str.upper() or " VS. " in rule_str.upper():
        parts = re.split(r'\s+VS\.?\s+', rule_str, flags=re.IGNORECASE)
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "AND"
    elif " & " in rule_str:
        parts = rule_str.split(" & ")
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "AND"
    elif "," in rule_str:
        parts = rule_str.split(",")
        ent1, ent2 = parts[0].strip(), parts[1].strip()
        op = "AND"
    else:
        ent1 = rule_str
        ent2 = rule_str
        op = "AND"

    return app_name, ent1 or "Permission 1", ent2 or "Permission 2", op


def normalize_imported_policy_item(item: dict) -> dict:
    """Normalizes any record dictionary (JSON or CSV/Excel row) into a standard SodPolicy format."""
    clean = {}
    for k, v in item.items():
        if k:
            clean_key = str(k).strip().lower().replace("_", " ").replace("-", " ")
            clean[clean_key] = v

    name = (
        clean.get("rule name") or
        clean.get("rulename") or
        clean.get("policy name") or
        clean.get("policyname") or
        clean.get("sod rule id") or
        clean.get("sodruleid") or
        clean.get("rule id") or
        clean.get("ruleid") or
        clean.get("sod id") or
        clean.get("name") or
        clean.get("title")
    )

    if not name:
        return None

    desc = clean.get("description") or clean.get("desc") or clean.get("details") or clean.get("summary") or f"SoD Policy {name}"
    
    risk_raw = str(clean.get("risk level") or clean.get("risklevel") or clean.get("risk") or clean.get("severity") or "HIGH").strip().upper()
    if "CRIT" in risk_raw:
        risk = "CRITICAL"
    elif "HIGH" in risk_raw:
        risk = "HIGH"
    elif "MED" in risk_raw:
        risk = "MEDIUM"
    elif "LOW" in risk_raw:
        risk = "LOW"
    else:
        risk = "HIGH"

    policy_type = str(clean.get("policy type") or clean.get("type") or "STATIC").strip().upper()
    status = str(clean.get("status") or "ACTIVE").strip().upper()
    business_owner = str(clean.get("business owner") or clean.get("owner") or "System").strip()
    approver = str(clean.get("approver") or "System").strip()

    # Extract rules
    rules_raw = (
        clean.get("rules") or
        clean.get("rule") or
        clean.get("conflicting entitlements") or
        clean.get("conflicting roles") or
        clean.get("entitlements") or
        clean.get("conflict")
    )

    parsed_rules = []

    if isinstance(rules_raw, list):
        for r in rules_raw:
            if isinstance(r, dict):
                r_clean = {str(rk).strip().lower().replace("_", " ").replace("-", " "): rv for rk, rv in r.items() if rk}
                app = r_clean.get("application name") or r_clean.get("application") or r_clean.get("app name") or r_clean.get("app") or "Global"
                e1 = r_clean.get("entitlement one") or r_clean.get("entitlement 1") or r_clean.get("entitlement1") or r_clean.get("role 1") or r_clean.get("permission 1") or ""
                e2 = r_clean.get("entitlement two") or r_clean.get("entitlement 2") or r_clean.get("entitlement2") or r_clean.get("role 2") or r_clean.get("permission 2") or ""
                cond = str(r_clean.get("condition type") or r_clean.get("condition") or r_clean.get("operator") or "AND").strip().upper()
                if e1 and e2:
                    parsed_rules.append({
                        "application_name": str(app),
                        "entitlement_one": str(e1),
                        "entitlement_two": str(e2),
                        "condition_type": cond if cond in ["AND", "OR", "NOT"] else "AND"
                    })
            elif isinstance(r, str) and r.strip():
                app, e1, e2, cond = parse_rule_expression_string(r)
                if e1 and e2:
                    parsed_rules.append({"application_name": app, "entitlement_one": e1, "entitlement_two": e2, "condition_type": cond})

    elif isinstance(rules_raw, str) and rules_raw.strip():
        app, e1, e2, cond = parse_rule_expression_string(rules_raw)
        if e1 and e2:
            parsed_rules.append({"application_name": app, "entitlement_one": e1, "entitlement_two": e2, "condition_type": cond})

    # Direct column check (e.g., Entitlement 1 & Entitlement 2 columns in the CSV/Excel row)
    if not parsed_rules:
        app = clean.get("application name") or clean.get("application") or clean.get("app name") or clean.get("app") or "Global"
        e1 = (
            clean.get("entitlement one") or clean.get("entitlement 1") or clean.get("entitlement1") or
            clean.get("first entitlement") or clean.get("role 1") or clean.get("role1") or
            clean.get("permission 1") or clean.get("function 1")
        )
        e2 = (
            clean.get("entitlement two") or clean.get("entitlement 2") or clean.get("entitlement2") or
            clean.get("second entitlement") or clean.get("role 2") or clean.get("role2") or
            clean.get("permission 2") or clean.get("function 2")
        )
        cond = str(clean.get("condition type") or clean.get("condition") or clean.get("operator") or "AND").strip().upper()
        if e1 and e2:
            parsed_rules.append({
                "application_name": str(app),
                "entitlement_one": str(e1),
                "entitlement_two": str(e2),
                "condition_type": cond if cond in ["AND", "OR", "NOT"] else "AND"
            })

    # Guaranteed fallback rule if entitlements could not be split
    if not parsed_rules:
        parsed_rules.append({
            "application_name": "Global",
            "entitlement_one": f"{name} Role A",
            "entitlement_two": f"{name} Role B",
            "condition_type": "AND"
        })

    return {
        "policy_name": str(name),
        "description": str(desc),
        "risk_level": risk,
        "policy_type": policy_type,
        "status": status,
        "business_owner": business_owner,
        "approver": approver,
        "rules": parsed_rules
    }


def parse_imported_policies_file(file: UploadFile) -> list:
    filename = file.filename.lower()
    content = file.file.read()
    
    raw_items = []

    if filename.endswith(".json"):
        raw = json.loads(content.decode("utf-8-sig", errors="ignore"))
        raw_items = raw if isinstance(raw, list) else [raw]
        
    elif filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        raw_items = list(reader)

    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        workbook = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or '').strip() for cell in sheet[1]]
        for row_cells in sheet.iter_rows(min_row=2, values_only=True):
            if any(row_cells):
                r_dict = {headers[i]: (row_cells[i] if i < len(headers) else None) for i in range(len(headers))}
                raw_items.append(r_dict)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload a .json, .csv, .xlsx, or .xls file."
        )

    # Group or normalize policies
    policies_map = {}
    for item in raw_items:
        normalized = normalize_imported_policy_item(item)
        if normalized:
            p_name = normalized["policy_name"]
            if p_name in policies_map:
                # Merge additional rules if same policy name appears on multiple rows
                policies_map[p_name]["rules"].extend(normalized["rules"])
            else:
                policies_map[p_name] = normalized

    return list(policies_map.values())


@router.post("/governance/sod-policies/import", dependencies=[Depends(require_permission("SoD Policies", "create"))])
def import_sod_policies(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    x_user_name: str = Header(default="System"),
    db: Session = Depends(get_db)
):
    """Imports policies from an uploaded JSON, CSV, or Excel (.xlsx/.xls) file and automatically triggers violation scans in background."""
    try:
        policies_list = parse_imported_policies_file(file)
        
        imported_count = 0
        skipped_count = 0
        
        for item in policies_list:
            name = item.get("policy_name")
            rules_data = item.get("rules", [])
            
            if not name or not rules_data:
                skipped_count += 1
                continue
                
            code = get_next_policy_code(db)
            policy = SodPolicy(
                policy_code=code,
                policy_name=name,
                description=item.get("description", ""),
                risk_level=item.get("risk_level", "HIGH"),
                policy_type=item.get("policy_type", "STATIC"),
                status=item.get("status", "ACTIVE"),
                business_owner=item.get("business_owner", "System"),
                approver=item.get("approver", "System"),
                created_by=x_user_name,
                version=1
            )
            db.add(policy)
            db.flush()
            
            for r in rules_data:
                rule = SodPolicyRule(
                    policy_id=policy.id,
                    application_name=r.get("application_name"),
                    entitlement_one=r.get("entitlement_one"),
                    entitlement_two=r.get("entitlement_two"),
                    condition_type=r.get("condition_type", "AND")
                )
                db.add(rule)
            
            db.commit()
            imported_count += 1
            write_sod_audit(db, policy.id, "Import", x_user_name, new_val={"policy_code": code, "policy_name": name})

        # --- Automatically queue SoD violation scan in background after policy import ---
        if imported_count > 0:
            try:
                from app.services.sod_violation_service import run_violation_scan_job
                scan = SodScanHistory(
                    scan_name=f"Auto Import Scan - {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
                    scan_type="FULL",
                    started_by=x_user_name,
                    status="RUNNING",
                    start_time=datetime.utcnow()
                )
                db.add(scan)
                db.commit()
                db.refresh(scan)
                background_tasks.add_task(run_violation_scan_job, db, scan.id, "FULL", x_user_name)
            except Exception as scan_err:
                print("Auto violation scan background queue warning:", scan_err)

        return {
            "imported": imported_count,
            "skipped": skipped_count,
            "auto_scanned": True,
            "message": f"Successfully imported {imported_count} policies from '{file.filename}'."
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse or save imported file: {e}")

# ── CSV & Excel Exports ──
@router.get("/governance/sod-policies/export/csv")
def export_csv(db: Session = Depends(get_db)):
    """Exports all policies to a CSV file."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Policy Code", "Policy Name", "Description", "Risk Level", 
        "Policy Type", "Status", "Business Owner", "Approver", "Created Date", "Rules Count"
    ])
    
    policies = db.query(SodPolicy).all()
    for p in policies:
        writer.writerow([
            p.policy_code,
            p.policy_name,
            p.description or "",
            p.risk_level,
            p.policy_type,
            p.status,
            p.business_owner,
            p.approver,
            p.created_date.strftime("%Y-%m-%d %H:%M:%S"),
            len(p.rules)
        ])
        
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sod_policies_export.csv"}
    )

@router.get("/governance/sod-policies/export/excel")
def export_excel(db: Session = Depends(get_db)):
    """Exports all policies to an Excel spreadsheet using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoD Policies"
    
    # Headers
    ws.append([
        "Policy Code", "Policy Name", "Description", "Risk Level", 
        "Policy Type", "Status", "Business Owner", "Approver", "Created Date", "Rules"
    ])
    
    policies = db.query(SodPolicy).all()
    for p in policies:
        rules_desc = "; ".join([
            f"[{r.application_name}]: {r.entitlement_one} {r.condition_type} {r.entitlement_two}"
            for r in p.rules
        ])
        ws.append([
            p.policy_code,
            p.policy_name,
            p.description or "",
            p.risk_level,
            p.policy_type,
            p.status,
            p.business_owner,
            p.approver,
            p.created_date.strftime("%Y-%m-%d %H:%M:%S"),
            rules_desc
        ])
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sod_policies_export.xlsx"}
    )
