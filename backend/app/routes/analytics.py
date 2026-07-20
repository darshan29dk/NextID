import csv
import io
import openpyxl

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.services.analytics_service import (
    get_executive_dashboard_data,
    get_role_analytics_data,
    get_coverage_report,
)

router = APIRouter()


@router.get("/analytics/executive")
def get_executive_dashboard(db: Session = Depends(get_db)):
    """AN-001: Executive Dashboard — platform-wide KPI dashboard."""
    return get_executive_dashboard_data(db)


@router.get("/analytics/role-analytics")
def get_role_analytics(db: Session = Depends(get_db)):
    """AN-002: Role Analytics — role-focused metrics."""
    return get_role_analytics_data(db)


@router.get("/analytics/coverage-reports")
def get_coverage_reports(db: Session = Depends(get_db)):
    """AN-003: Coverage Reports — identity/entitlement coverage breakdown."""
    return get_coverage_report(db)


# ── AN-004: Export Reports (CSV/Excel) ──────────────────────────────────
# Mirrors the export pattern already used in sod_policy.py (io.StringIO +
# csv.writer for CSV, openpyxl.Workbook for Excel, streamed back as a
# file-download response).

def _kpi_rows(kpis: dict):
    return [[k.replace("_", " ").title(), v] for k, v in kpis.items()]


@router.get("/analytics/executive/export/csv")
def export_executive_csv(db: Session = Depends(get_db)):
    data = get_executive_dashboard_data(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerows(_kpi_rows(data["kpis"]))
    for chart_name, chart_data in data["charts"].items():
        writer.writerow([])
        writer.writerow([chart_name.replace("_", " ").title()])
        for label, count in chart_data.items():
            writer.writerow([label, count])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=executive_dashboard_export.csv"}
    )


@router.get("/analytics/executive/export/excel")
def export_executive_excel(db: Session = Depends(get_db)):
    data = get_executive_dashboard_data(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"
    ws.append(["Metric", "Value"])
    for row in _kpi_rows(data["kpis"]):
        ws.append(row)
    for chart_name, chart_data in data["charts"].items():
        ws.append([])
        ws.append([chart_name.replace("_", " ").title()])
        for label, count in chart_data.items():
            ws.append([label, count])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=executive_dashboard_export.xlsx"}
    )


@router.get("/analytics/role-analytics/export/csv")
def export_role_analytics_csv(db: Session = Depends(get_db)):
    data = get_role_analytics_data(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerows(_kpi_rows(data["kpis"]))
    for chart_name, chart_data in data["charts"].items():
        writer.writerow([])
        writer.writerow([chart_name.replace("_", " ").title()])
        for label, count in chart_data.items():
            writer.writerow([label, count])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=role_analytics_export.csv"}
    )


@router.get("/analytics/role-analytics/export/excel")
def export_role_analytics_excel(db: Session = Depends(get_db)):
    data = get_role_analytics_data(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Role Analytics"
    ws.append(["Metric", "Value"])
    for row in _kpi_rows(data["kpis"]):
        ws.append(row)
    for chart_name, chart_data in data["charts"].items():
        ws.append([])
        ws.append([chart_name.replace("_", " ").title()])
        for label, count in chart_data.items():
            ws.append([label, count])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=role_analytics_export.xlsx"}
    )


@router.get("/analytics/coverage-reports/export/csv")
def export_coverage_csv(db: Session = Depends(get_db)):
    data = get_coverage_report(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerows(_kpi_rows(data["kpis"]))

    writer.writerow([])
    writer.writerow(["Coverage by Department"])
    writer.writerow(["Department", "Covered", "Total", "Coverage %"])
    for dept, stats in data["coverage_by_department"].items():
        writer.writerow([dept, stats["covered"], stats["total"], stats["coverage_pct"]])

    writer.writerow([])
    writer.writerow(["Uncovered Identities"])
    writer.writerow(["Name", "Email", "Department"])
    for u in data["uncovered_identities"]:
        writer.writerow([u["name"], u["email"], u["department"]])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=coverage_report_export.csv"}
    )


@router.get("/analytics/coverage-reports/export/excel")
def export_coverage_excel(db: Session = Depends(get_db)):
    data = get_coverage_report(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coverage Summary"
    ws.append(["Metric", "Value"])
    for row in _kpi_rows(data["kpis"]):
        ws.append(row)

    ws_dept = wb.create_sheet("By Department")
    ws_dept.append(["Department", "Covered", "Total", "Coverage %"])
    for dept, stats in data["coverage_by_department"].items():
        ws_dept.append([dept, stats["covered"], stats["total"], stats["coverage_pct"]])

    ws_unc = wb.create_sheet("Uncovered Identities")
    ws_unc.append(["Name", "Email", "Department"])
    for u in data["uncovered_identities"]:
        ws_unc.append([u["name"], u["email"], u["department"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=coverage_report_export.xlsx"}
    )
