from fastapi import APIRouter, Depends, HTTPException, status, Header, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_, asc, desc
from typing import List, Optional
import json
import os
import io
import re
import time
from datetime import datetime
import openpyxl

from app.database import get_db
from app.models.application import Application
from app.models.audit_log import AuditLog
from app.models.dashboard import RecentActivity
from app.models.platform_user import PlatformUser
from app.models.identity import Identity
from app.schemas.application import (
    ApplicationCreate, ApplicationUpdate, ApplicationResponse,
    ApplicationPaginatedResponse
)
from app.schemas.audit_log import AuditLogResponse
from app.models.application_account import ApplicationAccount
from app.models.import_run_history import ImportRunHistory
from app.models.application_entitlement import ApplicationEntitlement
from app.models.application_role import ApplicationRole
from app.models.application_field_mapping import ApplicationFieldMapping
from app.models.application_account_entitlement import ApplicationAccountEntitlement
from app.schemas.application_field_mapping import ApplicationFieldMappingItem, ApplicationFieldMappingResponse
router = APIRouter()

# Helper for Audit Logging
def write_application_audit(db: Session, user: str, action: str, old_val: dict = None, new_val: dict = None):
    try:
        old_val_str = json.dumps(old_val, default=str) if old_val else None
        new_val_str = json.dumps(new_val, default=str) if new_val else None

        audit = AuditLog(
            module="Applications",
            action=action,  # "Create", "Update", "Delete"
            performed_by=user,
            old_value=old_val_str,
            new_value=new_val_str,
            timestamp=datetime.utcnow()
        )
        db.add(audit)

        app_label = new_val.get("application_name") if new_val else (old_val.get("application_name") if old_val else "")
        activity = RecentActivity(
            user=user,
            action=f"Application {action.lower()}d - {app_label}",
            status="info" if action != "Delete" else "warning",
            created_at=datetime.utcnow()
        )
        db.add(activity)
        db.commit()
    except Exception as e:
        print(f"Warning: Failed to write application audit: {e}")


@router.get("/applications-owner-candidates")
def search_owner_candidates(
    q: Optional[str] = None,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """
    Returns candidate users (from PlatformUsers and Identities) to assign as Application Owner.
    """
    candidates = []
    seen_emails = set()

    pu_query = db.query(PlatformUser).filter(PlatformUser.is_deleted == False, PlatformUser.status == "Active")
    if q:
        like_term = f"%{q}%"
        pu_query = pu_query.filter(
            or_(
                PlatformUser.first_name.like(like_term),
                PlatformUser.last_name.like(like_term),
                PlatformUser.email.like(like_term),
                PlatformUser.employee_id.like(like_term),
                PlatformUser.department.like(like_term)
            )
        )
    p_users = pu_query.limit(limit).all()
    for u in p_users:
        full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username or u.email
        email = u.email or ""
        if email:
            seen_emails.add(email.lower())
        candidates.append({
            "id": u.id,
            "employee_id": u.employee_id or "",
            "name": full_name,
            "email": email,
            "source": "Platform User",
            "department": u.department or ""
        })

    if len(candidates) < limit:
        id_query = db.query(Identity).filter(Identity.is_deleted == False, Identity.status == "Active")
        if q:
            like_term = f"%{q}%"
            id_query = id_query.filter(
                or_(
                    Identity.display_name.like(like_term),
                    Identity.first_name.like(like_term),
                    Identity.last_name.like(like_term),
                    Identity.email.like(like_term),
                    Identity.employee_id.like(like_term),
                    Identity.department.like(like_term)
                )
            )
        identities = id_query.limit(limit).all()
        for i in identities:
            email = i.email or ""
            if email and email.lower() in seen_emails:
                continue
            full_name = i.display_name or f"{i.first_name or ''} {i.last_name or ''}".strip() or email
            if full_name:
                if email:
                    seen_emails.add(email.lower())
                candidates.append({
                    "id": i.id,
                    "employee_id": i.employee_id or "",
                    "name": full_name,
                    "email": email,
                    "source": "Identity",
                    "department": i.department or ""
                })

    return candidates[:limit]


@router.get("/applications", response_model=ApplicationPaginatedResponse)
def get_applications(
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    application_type: Optional[str] = None,
    status: Optional[str] = None,
    environment: Optional[str] = None,
    sortBy: Optional[str] = "created_at",
    sortOrder: Optional[str] = "desc",
    db: Session = Depends(get_db)
):
    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(Application).filter(Application.is_deleted == False)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Application.application_name.like(search_term),
                Application.description.like(search_term),
                Application.owner_name.like(search_term),
                Application.owner_email.like(search_term),
                Application.owner_employee_id.like(search_term)
            )
        )

    if application_type:
        query = query.filter(Application.application_type == application_type)
    if status:
        query = query.filter(Application.status == status)
    if environment:
        query = query.filter(Application.environment == environment)

    sort_fields = {
        "application_name": Application.application_name,
        "application_type": Application.application_type,
        "owner_name": Application.owner_name,
        "status": Application.status,
        "created_at": Application.created_at,
        "updated_at": Application.updated_at,
        "last_sync": Application.last_sync,
        "last_tested": Application.last_tested
    }

    sort_col = sort_fields.get(sortBy, Application.created_at)
    if sortOrder == "asc":
        query = query.order_by(asc(sort_col))
    else:
        query = query.order_by(desc(sort_col))

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    applications = query.offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "applications": applications
    }


@router.get("/applications/{id}", response_model=ApplicationResponse)
def get_application(id: int, db: Session = Depends(get_db)):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    return application


@router.post("/applications", response_model=ApplicationResponse)
def create_application(
    payload: ApplicationCreate,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    existing = db.query(Application).filter(
        Application.application_name == payload.application_name.strip(),
        Application.is_deleted == False
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="An application with this name already exists")

    if payload.application_type not in ["CSV", "Excel"]:
        raise HTTPException(status_code=400, detail="Application type must be 'CSV' or 'Excel'")

    application = Application(
        application_name=payload.application_name.strip(),
        application_type=payload.application_type,
        description=payload.description,
        status=payload.status or "Draft",
        health_status=payload.health_status or "Unknown",
        environment=payload.environment or "Development",
        tags=payload.tags,
        owner_id=payload.owner_id,
        owner_employee_id=payload.owner_employee_id.strip() if payload.owner_employee_id else None,
        owner_name=payload.owner_name.strip() if payload.owner_name else None,
        owner_email=payload.owner_email.strip() if payload.owner_email else None,
        csv_delimiter=payload.csv_delimiter or ",",
        csv_encoding=payload.csv_encoding or "UTF-8",
        excel_sheet_name=payload.excel_sheet_name,
        file_path=payload.file_path,
        created_by=x_user_name,
        modified_by=x_user_name
    )

    db.add(application)
    db.commit()
    db.refresh(application)

    app_dict = {
        "id": application.id,
        "application_name": application.application_name,
        "application_type": application.application_type,
        "status": application.status,
        "owner_employee_id": application.owner_employee_id,
        "owner_name": application.owner_name,
        "owner_email": application.owner_email
    }
    write_application_audit(db=db, user=x_user_name, action="Create", old_val=None, new_val=app_dict)

    return application


@router.put("/applications/{id}", response_model=ApplicationResponse)
def update_application(
    id: int,
    payload: ApplicationUpdate,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    old_dict = {
        "id": application.id,
        "application_name": application.application_name,
        "application_type": application.application_type,
        "status": application.status,
        "owner_employee_id": application.owner_employee_id,
        "owner_name": application.owner_name,
        "owner_email": application.owner_email
    }

    if payload.application_name and payload.application_name.strip() != application.application_name:
        existing = db.query(Application).filter(
            Application.application_name == payload.application_name.strip(),
            Application.is_deleted == False
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="An application with this name already exists")
        application.application_name = payload.application_name.strip()

    update_data = payload.dict(exclude_unset=True)
    update_data.pop("application_name", None)  # already handled above

    for key, value in update_data.items():
        if key in ["owner_employee_id", "owner_name", "owner_email"] and isinstance(value, str):
            value = value.strip() if value.strip() else None
        setattr(application, key, value)

    application.modified_by = x_user_name
    application.updated_at = datetime.utcnow()
    application.version = (application.version or 1) + 1

    db.commit()
    db.refresh(application)

    new_dict = {
        "id": application.id,
        "application_name": application.application_name,
        "application_type": application.application_type,
        "status": application.status,
        "owner_employee_id": application.owner_employee_id,
        "owner_name": application.owner_name,
        "owner_email": application.owner_email
    }
    write_application_audit(db=db, user=x_user_name, action="Update", old_val=old_dict, new_val=new_dict)

    return application


@router.delete("/applications/{id}")
def delete_application(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    old_dict = {
        "id": application.id,
        "application_name": application.application_name,
        "application_type": application.application_type,
        "status": application.status
    }

    application.is_deleted = True
    application.modified_by = x_user_name
    application.updated_at = datetime.utcnow()

    # Cascade: an Application's imported accounts and the entitlement
    # links on those accounts are meaningless once the parent Application
    # is gone, and previously stayed live in the DB even after delete -
    # inflating Analytics KPIs (Entitlements Mapped, etc.) with orphaned
    # rows from a source the user thought they'd removed. Entitlement
    # links have no is_deleted column of their own (they're a pure join
    # table), so they're hard-deleted here; accounts follow the same
    # soft-delete pattern used everywhere else in the app.
    accounts = db.query(ApplicationAccount).filter(
        ApplicationAccount.application_id == id, ApplicationAccount.is_deleted == False
    ).all()
    account_ids = [a.id for a in accounts]

    if account_ids:
        db.query(ApplicationAccountEntitlement).filter(
            ApplicationAccountEntitlement.application_id == id
        ).delete(synchronize_session=False)
        db.query(ApplicationAccount).filter(
            ApplicationAccount.id.in_(account_ids)
        ).update({ApplicationAccount.is_deleted: True, ApplicationAccount.modified_by: x_user_name}, synchronize_session=False)

    db.commit()

    write_application_audit(db=db, user=x_user_name, action="Delete", old_val=old_dict, new_val=None)

    return {"message": "Application deleted successfully"}


@router.post("/applications/{id}/upload")
async def upload_application_file(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    # If there is an existing file_path, delete the old file from disk
    if application.file_path:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        old_file_path = os.path.join(base_dir, application.file_path)
        if os.path.exists(old_file_path):
            try:
                os.remove(old_file_path)
            except Exception as e:
                print(f"Warning: Failed to delete old file {old_file_path}: {e}")

    content = await file.read()
    file_size = len(content)

    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    safe_filename = f"app_{application.id}_{file.filename}"
    file_path = os.path.join(uploads_dir, safe_filename)

    with open(file_path, "wb") as f:
        f.write(content)

    application.file_path = f"uploads/{safe_filename}"
    application.file_content = None  # Do not store raw file binary content in DB
    application.status = "Configured"
    application.modified_by = x_user_name
    application.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(application)

    return {
        "file_name": file.filename,
        "file_size": file_size,
        "file_path": application.file_path,
        "status": application.status
    }


@router.post("/applications/read-sheets")
async def read_excel_sheets(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported")
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheets = wb.sheetnames
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read sheets from Excel file: {str(e)}")


@router.get("/applications/{id}/audit-logs", response_model=List[AuditLogResponse])
def get_application_audit_logs(id: int, db: Session = Depends(get_db)):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    audits = db.query(AuditLog).filter(
        AuditLog.module == "Applications"
    ).order_by(AuditLog.timestamp.desc()).all()

    filtered_audits = []
    app_id_str = f'"id": {id}'
    app_name_str = f'"application_name": "{application.application_name}"'

    for a in audits:
        match = False
        if a.old_value and (app_id_str in a.old_value or app_name_str in a.old_value):
            match = True
        if a.new_value and (app_id_str in a.new_value or app_name_str in a.new_value):
            match = True
        if match:
            filtered_audits.append(a)

    return filtered_audits


@router.post("/applications/{id}/test")
def test_application(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, application)

    start_time = time.time()
    success = False
    message = ""

    try:
        if application.application_type == "CSV":
            if not application.file_path:
                raise Exception("No CSV file has been uploaded to this application yet.")

            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                application.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {application.file_path}")

            import csv as csv_module
            with open(full_path, "r", encoding=application.csv_encoding or "UTF-8") as f:
                reader = csv_module.reader(f, delimiter=application.csv_delimiter or ",")
                header = next(reader, None)
                if not header:
                    raise Exception("CSV file appears to be empty or has no header row.")
                message = f"Successfully read CSV file. Detected {len(header)} column(s): {', '.join(header[:5])}{'...' if len(header) > 5 else ''}"
            success = True

        elif application.application_type == "Excel":
            if not application.file_path:
                raise Exception("No Excel file has been uploaded to this application yet.")

            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                application.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {application.file_path}")

            wb = openpyxl.load_workbook(full_path, read_only=True)
            if application.excel_sheet_name and application.excel_sheet_name not in wb.sheetnames:
                raise Exception(f"Configured sheet '{application.excel_sheet_name}' not found in workbook. Available: {', '.join(wb.sheetnames)}")

            sheet = wb[application.excel_sheet_name] if application.excel_sheet_name else wb.active
            header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            col_count = len([c for c in header_row if c is not None]) if header_row else 0
            message = f"Successfully opened workbook sheet '{sheet.title}'. Detected {col_count} column(s)."
            success = True

    except Exception as e:
        success = False
        message = str(e)

    duration_ms = int((time.time() - start_time) * 1000)

    application.last_tested = datetime.utcnow()
    application.last_sync_duration = duration_ms
    if success:
        application.health_status = "Healthy"
        application.success_count = (application.success_count or 0) + 1
    else:
        application.health_status = "Unhealthy"
        application.failure_count = (application.failure_count or 0) + 1

    db.commit()

    return {
        "success": success,
        "message": message,
        "duration_ms": duration_ms,
        "tested_at": application.last_tested.isoformat()
    }


@router.get("/applications/{id}/schema")
def get_application_schema(id: int, db: Session = Depends(get_db)):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, application)

    fields = []

    try:
        if application.application_type == "CSV":
            if not application.file_path:
                raise Exception("No CSV file has been uploaded to this application yet.")
            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                application.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {application.file_path}")

            import csv as csv_module
            with open(full_path, "r", encoding=application.csv_encoding or "UTF-8") as f:
                reader = csv_module.reader(f, delimiter=application.csv_delimiter or ",")
                header = next(reader, None)
                if not header:
                    raise Exception("CSV file appears to be empty or has no header row.")
                sample_row = next(reader, None)
                for idx, col_name in enumerate(header):
                    sample_val = sample_row[idx] if sample_row and idx < len(sample_row) else None
                    fields.append({
                        "field_name": col_name.strip(),
                        "data_type": "String",
                        "sample_value": sample_val
                    })

        elif application.application_type == "Excel":
            if not application.file_path:
                raise Exception("No Excel file has been uploaded to this application yet.")
            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                application.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {application.file_path}")

            wb = openpyxl.load_workbook(full_path, read_only=True)
            sheet = wb[application.excel_sheet_name] if application.excel_sheet_name and application.excel_sheet_name in wb.sheetnames else wb.active

            rows_iter = sheet.iter_rows(max_row=2, values_only=True)
            header_row = next(rows_iter, None)
            sample_row = next(rows_iter, None)
            if not header_row:
                raise Exception("Excel sheet appears to be empty.")

            for idx, col_name in enumerate(header_row):
                if col_name is None:
                    continue
                sample_val = sample_row[idx] if sample_row and idx < len(sample_row) else None
                fields.append({
                    "field_name": str(col_name).strip(),
                    "data_type": "String",
                    "sample_value": str(sample_val) if sample_val is not None else None
                })

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {"fields": fields, "field_count": len(fields)}
def _read_all_rows(application):
    """Reads every data row from the application's uploaded file as a list of dicts."""
    if not application.file_path:
        raise Exception("No file has been uploaded to this application yet.")
    full_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        application.file_path
    )
    if not os.path.exists(full_path):
        raise Exception(f"File not found on server: {application.file_path}")

    rows = []
    if application.application_type == "CSV":
        import csv as csv_module
        with open(full_path, "r", encoding=application.csv_encoding or "UTF-8") as f:
            reader = csv_module.DictReader(f, delimiter=application.csv_delimiter or ",")
            for row in reader:
                rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k})
    elif application.application_type == "Excel":
        wb = openpyxl.load_workbook(full_path, read_only=True)
        sheet = wb[application.excel_sheet_name] if application.excel_sheet_name and application.excel_sheet_name in wb.sheetnames else wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise Exception("Excel sheet appears to be empty.")
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        for row in rows_iter:
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = val
            if any(v is not None for v in row_dict.values()):
                rows.append(row_dict)
    else:
        raise Exception(f"Import not supported for application type: {application.application_type}")
    return rows


def _find_field(row_dict, candidates):
    """Case-insensitive match of a row's keys against a list of likely field names."""
    lower_map = {str(k).lower().strip(): v for k, v in row_dict.items()}
    for cand in candidates:
        if cand in lower_map:
            val = lower_map[cand]
            return str(val) if val is not None else None
    return None


def _clean_row_for_json(row):
    return {str(k): (str(v) if v is not None else None) for k, v in row.items()}


def _get_field_mapping_dict(db, application_id, target_module):
    """Returns {target_attribute_name: source_field} for this application's saved mappings."""
    mappings = db.query(ApplicationFieldMapping).filter(
        ApplicationFieldMapping.application_id == application_id,
        ApplicationFieldMapping.target_module == target_module
    ).all()
    return {m.target_attribute_name: m.source_field for m in mappings}


def _resolve_field(row, mapping_dict, target_attr, fallback_candidates):
    """Priority: user's saved mapping for this attribute > built-in guess list."""
    if target_attr in mapping_dict:
        mapped_source = mapping_dict[target_attr]
        lower_map = {str(k).lower().strip(): v for k, v in row.items()}
        val = lower_map.get(str(mapped_source).lower().strip())
        if val is not None:
            return str(val)
    return _find_field(row, fallback_candidates)

@router.post("/applications/{id}/import-accounts")
def import_accounts(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, application)

    start_time = time.time()
    try:
        rows = _read_all_rows(application)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rows:
        return {"success": True, "total": 0, "imported": 0, "errors": 0, "duration_ms": 0}

    # Do not wipe all accounts. Instead, upsert each row.
    mapping_dict = _get_field_mapping_dict(db, id, "Account")

    # Build a case-insensitive lookup of entitlements already imported for this
    # application, so account rows carrying an "entitlements" column can be
    # linked to them by name.
    existing_entitlements = db.query(ApplicationEntitlement).filter(
        ApplicationEntitlement.application_id == id,
        ApplicationEntitlement.is_deleted == False
    ).all()
    entitlement_lookup = {e.entitlement_name.strip().lower(): e for e in existing_entitlements if e.entitlement_name}

    imported_count = 0
    error_count = 0
    entitlement_links_created = 0
    unmatched_entitlement_names = set()

    # Load existing accounts in a single query to avoid N+1 query overhead over remote DB
    existing_accounts = db.query(ApplicationAccount).filter(
        ApplicationAccount.application_id == id,
        ApplicationAccount.is_deleted == False
    ).all()
    account_lookup = {acc.account_id: acc for acc in existing_accounts}

    for idx, row in enumerate(rows):
        try:
            account_id_val = _resolve_field(row, mapping_dict, "account_id", ["account_id", "id", "employee_id", "user_id", "emp_id"]) or f"row_{idx + 1}"
            account_name_val = _resolve_field(row, mapping_dict, "account_name", ["account_name", "name", "username", "full_name"])
            email_val = _resolve_field(row, mapping_dict, "email", ["email", "email_address"])
            status_val = _resolve_field(row, mapping_dict, "status", ["status", "active_status"]) or "Active"
            entitlements_val = _resolve_field(row, mapping_dict, "entitlements", ["entitlements", "entitlement", "groups", "group", "roles", "permissions"])

            existing_account = account_lookup.get(account_id_val)

            if existing_account:
                # Update existing account
                existing_account.account_name = account_name_val
                existing_account.email = email_val
                existing_account.status = status_val
                existing_account.raw_data = _clean_row_for_json(row)
                existing_account.modified_by = x_user_name
                existing_account.updated_at = datetime.utcnow()

                # Delete old links for this account
                db.query(ApplicationAccountEntitlement).filter(
                    ApplicationAccountEntitlement.application_id == id,
                    ApplicationAccountEntitlement.account_id == existing_account.id
                ).delete()

                record = existing_account
            else:
                # Insert new account
                record = ApplicationAccount(
                    application_id=id,
                    account_id=account_id_val,
                    account_name=account_name_val,
                    email=email_val,
                    status=status_val,
                    raw_data=_clean_row_for_json(row),
                    created_by=x_user_name,
                    modified_by=x_user_name
                )
                db.add(record)
                imported_count += 1

            if entitlements_val:
                db.flush()  # assign record.id so the link rows can reference it
                names = [n.strip() for n in re.split(r"[;,]", entitlements_val) if n.strip()]
                for name in names:
                    matched_entitlement = entitlement_lookup.get(name.lower())
                    link = ApplicationAccountEntitlement(
                        application_id=id,
                        account_id=record.id,
                        entitlement_id=matched_entitlement.id if matched_entitlement else None,
                        entitlement_name_raw=name,
                        matched=matched_entitlement is not None
                    )
                    db.add(link)
                    entitlement_links_created += 1
                    if not matched_entitlement:
                        unmatched_entitlement_names.add(name)
        except Exception:
            error_count += 1

    db.commit()
    duration_ms = int((time.time() - start_time) * 1000)

    history = ImportRunHistory(
        source_type="Application",
        source_id=id,
        run_type="Account Import",
        total_records=len(rows),
        valid_records=imported_count,
        warning_records=len(unmatched_entitlement_names),
        error_records=error_count,
        status="Completed" if error_count == 0 else "Partial",
        run_by=x_user_name
    )
    db.add(history)

    application.last_sync = datetime.utcnow()
    application.last_sync_duration = duration_ms
    application.success_count = (application.success_count or 0) + 1
    db.commit()

    write_application_audit(
        db=db, user=x_user_name, action="Import Accounts",
        old_val=None,
        new_val={"id": application.id, "application_name": application.application_name, "imported": imported_count, "errors": error_count}
    )

    return {
        "success": True,
        "total": len(rows),
        "imported": imported_count,
        "errors": error_count,
        "duration_ms": duration_ms,
        "entitlement_assignments_imported": entitlement_links_created,
        "unmatched_entitlement_names": sorted(unmatched_entitlement_names)
    }


@router.get("/applications/{id}/accounts")
def get_application_accounts(
    id: int,
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(ApplicationAccount).filter(
        ApplicationAccount.application_id == id,
        ApplicationAccount.is_deleted == False
    )
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                ApplicationAccount.account_id.like(search_term),
                ApplicationAccount.account_name.like(search_term),
                ApplicationAccount.email.like(search_term)
            )
        )

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    accounts = query.order_by(ApplicationAccount.id.asc()).offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "accounts": [
            {
                "id": a.id,
                "account_id": a.account_id,
                "account_name": a.account_name,
                "email": a.email,
                "status": a.status,
                "imported_at": a.imported_at.isoformat() if a.imported_at else None,
                "raw_data": a.raw_data
            } for a in accounts
        ]
    }
@router.post("/applications/{id}/import-entitlements")
def import_entitlements(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, application)

    start_time = time.time()
    try:
        rows = _read_all_rows(application)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rows:
        return {"success": True, "total": 0, "imported": 0, "errors": 0, "duration_ms": 0}

    # Do not wipe all entitlements. Instead, upsert each row.
    mapping_dict = _get_field_mapping_dict(db, id, "Entitlement")

    # Load existing entitlements in a single query to avoid N+1 query overhead over remote DB
    existing_ents = db.query(ApplicationEntitlement).filter(
        ApplicationEntitlement.application_id == id,
        ApplicationEntitlement.is_deleted == False
    ).all()
    ent_lookup = {ent.entitlement_name.lower(): ent for ent in existing_ents if ent.entitlement_name}

    imported_count = 0
    error_count = 0
    for idx, row in enumerate(rows):
        try:
            name_val = _resolve_field(row, mapping_dict, "entitlement_name", ["entitlement_name", "name", "entitlement", "permission_name", "access_name"]) or f"row_{idx + 1}"
            type_val = _resolve_field(row, mapping_dict, "entitlement_type", ["entitlement_type", "type", "category", "permission_type"])
            desc_val = _resolve_field(row, mapping_dict, "description", ["description", "desc"])

            existing_entitlement = ent_lookup.get(name_val.lower())

            if existing_entitlement:
                # Update existing entitlement
                existing_entitlement.entitlement_type = type_val
                existing_entitlement.description = desc_val
                existing_entitlement.raw_data = _clean_row_for_json(row)
                existing_entitlement.modified_by = x_user_name
                existing_entitlement.updated_at = datetime.utcnow()
            else:
                # Insert new entitlement
                record = ApplicationEntitlement(
                    application_id=id,
                    entitlement_name=name_val,
                    entitlement_type=type_val,
                    description=desc_val,
                    raw_data=_clean_row_for_json(row),
                    created_by=x_user_name,
                    modified_by=x_user_name
                )
                db.add(record)
                imported_count += 1
        except Exception:
            error_count += 1

    db.commit()
    duration_ms = int((time.time() - start_time) * 1000)

    history = ImportRunHistory(
        source_type="Application",
        source_id=id,
        run_type="Entitlement Import",
        total_records=len(rows),
        valid_records=imported_count,
        warning_records=0,
        error_records=error_count,
        status="Completed" if error_count == 0 else "Partial",
        run_by=x_user_name
    )
    db.add(history)

    application.last_sync = datetime.utcnow()
    application.last_sync_duration = duration_ms
    application.success_count = (application.success_count or 0) + 1
    db.commit()

    write_application_audit(
        db=db, user=x_user_name, action="Import Entitlements",
        old_val=None,
        new_val={"id": application.id, "application_name": application.application_name, "imported": imported_count, "errors": error_count}
    )

    return {
        "success": True,
        "total": len(rows),
        "imported": imported_count,
        "errors": error_count,
        "duration_ms": duration_ms
    }


@router.get("/applications/{id}/entitlements")
def get_application_entitlements(
    id: int,
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(ApplicationEntitlement).filter(
        ApplicationEntitlement.application_id == id,
        ApplicationEntitlement.is_deleted == False
    )
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                ApplicationEntitlement.entitlement_name.like(search_term),
                ApplicationEntitlement.entitlement_type.like(search_term)
            )
        )

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    entitlements = query.order_by(ApplicationEntitlement.id.asc()).offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "entitlements": [
            {
                "id": e.id,
                "entitlement_name": e.entitlement_name,
                "entitlement_type": e.entitlement_type,
                "description": e.description,
                "imported_at": e.imported_at.isoformat() if e.imported_at else None,
                "raw_data": e.raw_data
            } for e in entitlements
        ]
    }


@router.post("/applications/{id}/import-roles")
def import_roles(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, application)

    start_time = time.time()
    try:
        rows = _read_all_rows(application)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rows:
        return {"success": True, "total": 0, "imported": 0, "errors": 0, "duration_ms": 0}

    # Do not wipe all roles. Instead, upsert each row.
    mapping_dict = _get_field_mapping_dict(db, id, "Role")

    # Load existing roles in a single query to avoid N+1 query overhead over remote DB
    existing_roles = db.query(ApplicationRole).filter(
        ApplicationRole.application_id == id,
        ApplicationRole.is_deleted == False
    ).all()
    role_lookup = {r.role_name.lower(): r for r in existing_roles if r.role_name}

    imported_count = 0
    error_count = 0
    for idx, row in enumerate(rows):
        try:
            name_val = _resolve_field(row, mapping_dict, "role_name", ["role_name", "name", "role"]) or f"row_{idx + 1}"
            desc_val = _resolve_field(row, mapping_dict, "description", ["description", "desc"])

            existing_role = role_lookup.get(name_val.lower())

            if existing_role:
                # Update existing role
                existing_role.description = desc_val
                existing_role.raw_data = _clean_row_for_json(row)
                existing_role.modified_by = x_user_name
                existing_role.updated_at = datetime.utcnow()
            else:
                # Insert new role
                record = ApplicationRole(
                    application_id=id,
                    role_name=name_val,
                    description=desc_val,
                    raw_data=_clean_row_for_json(row),
                    created_by=x_user_name,
                    modified_by=x_user_name
                )
                db.add(record)
                imported_count += 1
        except Exception:
            error_count += 1

    db.commit()
    duration_ms = int((time.time() - start_time) * 1000)

    history = ImportRunHistory(
        source_type="Application",
        source_id=id,
        run_type="Role Import",
        total_records=len(rows),
        valid_records=imported_count,
        warning_records=0,
        error_records=error_count,
        status="Completed" if error_count == 0 else "Partial",
        run_by=x_user_name
    )
    db.add(history)

    application.last_sync = datetime.utcnow()
    application.last_sync_duration = duration_ms
    application.success_count = (application.success_count or 0) + 1
    db.commit()

    write_application_audit(
        db=db, user=x_user_name, action="Import Roles",
        old_val=None,
        new_val={"id": application.id, "application_name": application.application_name, "imported": imported_count, "errors": error_count}
    )

    return {
        "success": True,
        "total": len(rows),
        "imported": imported_count,
        "errors": error_count,
        "duration_ms": duration_ms
    }


@router.get("/applications/{id}/roles")
def get_application_roles(
    id: int,
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(ApplicationRole).filter(
        ApplicationRole.application_id == id,
        ApplicationRole.is_deleted == False
    )
    if search:
        search_term = f"%{search}%"
        query = query.filter(ApplicationRole.role_name.like(search_term))

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    roles = query.order_by(ApplicationRole.id.asc()).offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "roles": [
            {
                "id": r.id,
                "role_name": r.role_name,
                "description": r.description,
                "imported_at": r.imported_at.isoformat() if r.imported_at else None,
                "raw_data": r.raw_data
            } for r in roles
        ]
    }
@router.get("/applications/{id}/mappings", response_model=List[ApplicationFieldMappingResponse])
def get_application_mappings(id: int, db: Session = Depends(get_db)):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    mappings = db.query(ApplicationFieldMapping).filter(ApplicationFieldMapping.application_id == id).all()
    return mappings


@router.put("/applications/{id}/mappings")
def save_application_mappings(
    id: int,
    payload: List[ApplicationFieldMappingItem],
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    db.query(ApplicationFieldMapping).filter(ApplicationFieldMapping.application_id == id).delete()

    for item in payload:
        mapping = ApplicationFieldMapping(
            application_id=id,
            source_field=item.source_field,
            target_module=item.target_module,
            target_attribute_name=item.target_attribute_name,
            transformation_type=item.transformation_type,
            created_by=x_user_name,
            modified_by=x_user_name
        )
        db.add(mapping)

    db.commit()

    write_application_audit(
        db=db, user=x_user_name, action="Update Mapping",
        old_val=None,
        new_val={"id": application.id, "application_name": application.application_name, "mapping_count": len(payload)}
    )

    return {"message": "Mappings saved successfully", "count": len(payload)}
@router.get("/applications/{id}/import-history")
def get_application_import_history(
    id: int,
    page: int = 1,
    limit: int = 25,
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == id, Application.is_deleted == False).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(ImportRunHistory).filter(
        ImportRunHistory.source_type == "Application",
        ImportRunHistory.source_id == id
    ).order_by(ImportRunHistory.run_at.desc())

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    history = query.offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "history": [
            {
                "id": h.id,
                "run_type": h.run_type,
                "total_records": h.total_records,
                "valid_records": h.valid_records,
                "warning_records": h.warning_records,
                "error_records": h.error_records,
                "status": h.status,
                "run_by": h.run_by,
                "run_at": h.run_at.isoformat() if h.run_at else None
            } for h in history
        ]
    }