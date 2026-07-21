from fastapi import APIRouter, Depends, HTTPException, status, Header, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, asc, desc
from sqlalchemy.exc import IntegrityError
from typing import List, Optional
import json
import os
import io
import time
from datetime import datetime
import openpyxl

from app.database import get_db
from app.models.connector import Connector
from app.models.connector_log import ConnectorLog
from app.models.connector_file import ConnectorFile
from app.models.audit_log import AuditLog
from app.models.dashboard import RecentActivity
from app.schemas.connector import (
    ConnectorCreate, ConnectorUpdate, ConnectorResponse, 
    ConnectorPaginatedResponse, ConnectorLogResponse, ConnectorFileResponse
)
from app.schemas.audit_log import AuditLogResponse
from app.utils.crypto import encrypt_password, decrypt_password

router = APIRouter()

# Helper for Audit Logging
def write_connector_audit(db: Session, user: str, action: str, old_val: dict = None, new_val: dict = None):
    try:
        old_val_str = json.dumps(old_val, default=str) if old_val else None
        new_val_str = json.dumps(new_val, default=str) if new_val else None

        audit = AuditLog(
            module="Connectors",
            action=action, # "Create", "Update", "Delete"
            performed_by=user,
            old_value=old_val_str,
            new_value=new_val_str,
            timestamp=datetime.utcnow()
        )
        db.add(audit)

        # Recent Activity Feed
        conn_label = new_val.get("connector_name") if new_val else (old_val.get("connector_name") if old_val else "")
        activity = RecentActivity(
            user=user,
            action=f"Connector {action.lower()}d - {conn_label}",
            status="info" if action != "Delete" else "warning",
            created_at=datetime.utcnow()
        )
        db.add(activity)
        db.commit()
    except Exception as e:
        print(f"Warning: Failed to write connector audit: {e}")

# Helper for Connector Logs
def write_connector_log(db: Session, connector_id: int, action: str, details: str, status_val: str):
    try:
        log = ConnectorLog(
            connector_id=connector_id,
            action=action,
            details=details,
            status=status_val,
            timestamp=datetime.utcnow()
        )
        db.add(log)
        db.commit()
    except Exception as e:
        print(f"Warning: Failed to write connector log: {e}")

@router.get("/connectors", response_model=ConnectorPaginatedResponse)
def get_connectors(
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    connector_type: Optional[str] = None,
    status: Optional[str] = None,
    database_type: Optional[str] = None,
    sortBy: Optional[str] = "created_at",
    sortOrder: Optional[str] = "desc",
    db: Session = Depends(get_db)
):
    if page < 1:
        page = 1
    if limit < 1:
        limit = 25

    query = db.query(Connector).filter(Connector.is_deleted == False)

    # 1. Search Query
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Connector.connector_name.like(search_term),
                Connector.description.like(search_term),
                Connector.host.like(search_term),
                Connector.database_name.like(search_term)
            )
        )

    # 2. Filters
    if connector_type:
        query = query.filter(Connector.connector_type == connector_type)
    if status:
        query = query.filter(Connector.status == status)
    if database_type:
        query = query.filter(Connector.database_type == database_type)

    # 3. Sorting
    sort_fields = {
        "connector_name": Connector.connector_name,
        "connector_type": Connector.connector_type,
        "status": Connector.status,
        "database_type": Connector.database_type,
        "created_at": Connector.created_at,
        "updated_at": Connector.updated_at,
        "last_sync": Connector.last_sync,
        "last_tested": Connector.last_tested
    }

    sort_col = sort_fields.get(sortBy, Connector.created_at)
    if sortOrder == "asc":
        query = query.order_by(asc(sort_col))
    else:
        query = query.order_by(desc(sort_col))

    total = query.count()
    total_pages = (total + limit - 1) // limit if total > 0 else 0
    offset = (page - 1) * limit
    connectors = query.offset(offset).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "connectors": connectors
    }

@router.get("/connectors/{id}", response_model=ConnectorResponse)
def get_connector(id: int, db: Session = Depends(get_db)):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")
    return connector

@router.post("/connectors", response_model=ConnectorResponse)
def create_connector(
    payload: ConnectorCreate,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    if not payload.connector_name or not payload.connector_name.strip():
        raise HTTPException(status_code=400, detail="Connector name is required")

    name = payload.connector_name.strip()
    # Unique constraint check
    existing = db.query(Connector).filter(
        Connector.connector_name == name,
        Connector.is_deleted == False
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"A connector named '{name}' already exists")

    encrypted_pw = None
    if payload.password:
        encrypted_pw = encrypt_password(payload.password)

    connector = Connector(
        connector_name=name,
        connector_type=payload.connector_type,
        description=payload.description,
        status=payload.status or "Draft",
        health_status=payload.health_status or "Unknown",
        environment=payload.environment or "Development",
        auth_type=payload.auth_type or "Basic",
        tags=payload.tags,
        database_type=payload.database_type,
        host=payload.host,
        port=payload.port,
        database_name=payload.database_name,
        username=payload.username,
        password=encrypted_pw,
        ssl_enabled=payload.ssl_enabled or False,
        connection_timeout=payload.connection_timeout or 30,
        csv_delimiter=payload.csv_delimiter or ",",
        csv_encoding=payload.csv_encoding or "UTF-8",
        excel_sheet_name=payload.excel_sheet_name,
        file_path=payload.file_path,
        created_by=x_user_name,
        modified_by=x_user_name
    )

    db.add(connector)
    try:
        db.commit()
        db.refresh(connector)
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"A connector named '{name}' already exists"
        )

    # Logging and auditing
    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Created",
        details=f"Connector '{connector.connector_name}' successfully configured as {connector.connector_type}.",
        status_val="Success"
    )

    conn_dict = {
        "id": connector.id,
        "connector_name": connector.connector_name,
        "connector_type": connector.connector_type,
        "status": connector.status
    }
    write_connector_audit(db=db, user=x_user_name, action="Create", old_val=None, new_val=conn_dict)

    return connector

@router.put("/connectors/{id}", response_model=ConnectorResponse)
def update_connector(
    id: int,
    payload: ConnectorUpdate,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    old_dict = {
        "id": connector.id,
        "connector_name": connector.connector_name,
        "connector_type": connector.connector_type,
        "status": connector.status
    }

    # Update logic
    update_data = payload.dict(exclude_unset=True)

    if "connector_name" in update_data and update_data["connector_name"] is not None:
        stripped_name = update_data["connector_name"].strip()
        if not stripped_name:
            raise HTTPException(status_code=400, detail="Connector name cannot be empty")
        update_data["connector_name"] = stripped_name
        if stripped_name != connector.connector_name:
            existing = db.query(Connector).filter(
                Connector.connector_name == stripped_name,
                Connector.is_deleted == False
            ).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"A connector named '{stripped_name}' already exists")
    
    # Handle password encryption update specifically
    if "password" in update_data:
        if update_data["password"]:
            connector.password = encrypt_password(update_data["password"])
        else:
            connector.password = None
        del update_data["password"]

    for key, value in update_data.items():
        setattr(connector, key, value)

    connector.modified_by = x_user_name
    connector.updated_at = datetime.utcnow()

    # Bump version
    connector.version = (connector.version or 1) + 1

    try:
        db.commit()
        db.refresh(connector)
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="A connector with this name already exists"
        )

    # Logging and auditing
    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Updated",
        details=f"Connector configuration updated by {x_user_name}.",
        status_val="Success"
    )

    new_dict = {
        "id": connector.id,
        "connector_name": connector.connector_name,
        "connector_type": connector.connector_type,
        "status": connector.status
    }
    write_connector_audit(db=db, user=x_user_name, action="Update", old_val=old_dict, new_val=new_dict)

    return connector

@router.delete("/connectors/{id}")
def delete_connector(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    old_dict = {
        "id": connector.id,
        "connector_name": connector.connector_name,
        "connector_type": connector.connector_type,
        "status": connector.status
    }

    # Soft delete
    connector.is_deleted = True
    if not "_deleted_" in connector.connector_name:
        connector.connector_name = f"{connector.connector_name}_deleted_{connector.id}"
    connector.modified_by = x_user_name
    connector.updated_at = datetime.utcnow()

    db.commit()

    write_connector_log(
        db=db,
        connector_id=id,
        action="Deleted",
        details=f"Connector soft deleted by {x_user_name}.",
        status_val="Success"
    )

    write_connector_audit(db=db, user=x_user_name, action="Delete", old_val=old_dict, new_val=None)

    return {"message": "Connector deleted successfully"}

@router.post("/connectors/{id}/upload", response_model=ConnectorFileResponse)
async def upload_connector_file(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    # If there is an existing file_path, delete the old file from disk
    if connector.file_path:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        old_file_path = os.path.join(base_dir, connector.file_path)
        if os.path.exists(old_file_path):
            try:
                os.remove(old_file_path)
            except Exception as e:
                print(f"Warning: Failed to delete old file {old_file_path}: {e}")

    # Read and save new file content
    content = await file.read()
    file_size = len(content)

    # Ensure uploads directory exists inside backend/
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    safe_filename = f"{connector.id}_{file.filename}"
    file_path = os.path.join(uploads_dir, safe_filename)

    with open(file_path, "wb") as f:
        f.write(content)

    # Delete old ConnectorFile records for this connector from DB
    try:
        db.query(ConnectorFile).filter(ConnectorFile.connector_id == id).delete(synchronize_session=False)
    except Exception as e:
        print(f"Warning: Failed to delete old ConnectorFile records: {e}")

    # Create file record (do not store raw file binary content in DB)
    conn_file = ConnectorFile(
        connector_id=id,
        file_name=file.filename,
        file_type=file.content_type or "application/octet-stream",
        file_size=file_size,
        uploaded_by=x_user_name,
        upload_date=datetime.utcnow(),
        file_content=None
    )

    db.add(conn_file)

    # Update connector file path config
    connector.file_path = f"uploads/{safe_filename}"
    connector.status = "Configured"
    db.commit()
    db.refresh(conn_file)

    write_connector_log(
        db=db,
        connector_id=id,
        action="Import Started",
        details=f"File '{file.filename}' uploaded successfully. Connector status shifted to Configured.",
        status_val="Success"
    )

    return conn_file

@router.post("/connectors/read-sheets")
async def read_excel_sheets(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported")
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheets = wb.sheetnames
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read sheets from Excel file: {str(e)}")

@router.get("/connectors/{id}/logs", response_model=List[ConnectorLogResponse])
def get_connector_logs(id: int, db: Session = Depends(get_db)):
    # Verify connector exists
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")
        
    logs = db.query(ConnectorLog).filter(ConnectorLog.connector_id == id).order_by(ConnectorLog.timestamp.desc()).all()
    return logs

@router.get("/connectors/{id}/files", response_model=List[ConnectorFileResponse])
def get_connector_files(id: int, db: Session = Depends(get_db)):
    # Verify connector exists
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")
        
    files = db.query(ConnectorFile).filter(ConnectorFile.connector_id == id).order_by(ConnectorFile.upload_date.desc()).all()
    return files

@router.get("/connectors/{id}/audit-logs", response_model=List[AuditLogResponse])
def get_connector_audit_logs(id: int, db: Session = Depends(get_db)):
    # Verify connector exists
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")
        
    # Query audits where module is "Connectors" and the log matches this connector's ID or name
    audits = db.query(AuditLog).filter(
        AuditLog.module == "Connectors"
    ).order_by(AuditLog.timestamp.desc()).all()

    # Filter audits matching connector name or ID
    filtered_audits = []
    conn_id_str = f'"id": {id}'
    conn_name_str = f'"connector_name": "{connector.connector_name}"'
    
    for a in audits:
        match = False
        if a.old_value and (conn_id_str in a.old_value or conn_name_str in a.old_value):
            match = True
        if a.new_value and (conn_id_str in a.new_value or conn_name_str in a.new_value):
            match = True
        if match:
            filtered_audits.append(a)
            
    return filtered_audits

@router.post("/connectors/{id}/test")
def test_connector(
    id: int,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, connector)

    start_time = time.time()
    success = False
    message = ""

    try:
        if connector.connector_type == "CSV":
            if not connector.file_path:
                raise Exception("No CSV file has been uploaded to this connector yet.")

            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                connector.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {connector.file_path}")

            import csv as csv_module
            with open(full_path, "r", encoding=connector.csv_encoding or "UTF-8") as f:
                reader = csv_module.reader(f, delimiter=connector.csv_delimiter or ",")
                header = next(reader, None)
                if not header:
                    raise Exception("CSV file appears to be empty or has no header row.")
                message = f"Successfully read CSV file. Detected {len(header)} column(s): {', '.join(header[:5])}{'...' if len(header) > 5 else ''}"
            success = True

        elif connector.connector_type == "Excel":
            if not connector.file_path:
                raise Exception("No Excel file has been uploaded to this connector yet.")

            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                connector.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {connector.file_path}")

            wb = openpyxl.load_workbook(full_path, read_only=True)
            if connector.excel_sheet_name and connector.excel_sheet_name not in wb.sheetnames:
                raise Exception(f"Configured sheet '{connector.excel_sheet_name}' not found in workbook. Available: {', '.join(wb.sheetnames)}")

            sheet = wb[connector.excel_sheet_name] if connector.excel_sheet_name else wb.active
            header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            col_count = len([c for c in header_row if c is not None]) if header_row else 0
            message = f"Successfully opened workbook sheet '{sheet.title}'. Detected {col_count} column(s)."
            success = True

        elif connector.connector_type == "Database":
            if connector.database_type != "MySQL":
                raise Exception(f"Live connection testing for {connector.database_type} is not yet supported. Only MySQL is currently testable.")

            import pymysql
            decrypted_pw = decrypt_password(connector.password) if connector.password else ""

            conn = pymysql.connect(
                host=connector.host,
                port=connector.port or 3306,
                user=connector.username,
                password=decrypted_pw,
                database=connector.database_name,
                connect_timeout=connector.connection_timeout or 30
            )
            cursor = conn.cursor()
            cursor.execute("SELECT VERSION()")
            version = cursor.fetchone()
            cursor.close()
            conn.close()
            message = f"Successfully connected to MySQL database '{connector.database_name}'. Server version: {version[0] if version else 'Unknown'}"
            success = True

        elif connector.connector_type == "LDAP":
            raise Exception("LDAP connection testing is not yet implemented.")

        elif connector.connector_type == "API Gateway":
            import requests as requests_lib
            url = connector.host
            if not url:
                raise Exception("API Gateway URL is required.")
            
            headers = {}
            if connector.file_path:
                try:
                    headers = json.loads(connector.file_path)
                except Exception:
                    pass

            auth = None
            if connector.auth_type == "Basic":
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                auth = (connector.username, decrypted_pw)
            elif connector.auth_type == "API Key" and connector.username and connector.password:
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                headers[connector.username] = decrypted_pw

            timeout = connector.connection_timeout or 30
            resp = requests_lib.get(url, headers=headers, auth=auth, timeout=timeout)
            
            if resp.status_code >= 400:
                raise Exception(f"HTTP Error {resp.status_code}: {resp.text[:100]}")
            
            res_json = resp.json()
            json_key = connector.database_name
            target_list = res_json
            if json_key:
                if json_key in res_json:
                    target_list = res_json[json_key]
            
            if isinstance(target_list, list):
                message = f"Successfully connected to API. Extracted {len(target_list)} records from path '{json_key or 'root'}'."
            elif isinstance(res_json, dict):
                message = f"Successfully connected to API. Received JSON response object."
            else:
                message = f"Successfully connected to API. Response was text."
            
            success = True

    except Exception as e:
        success = False
        message = str(e)

    duration_ms = int((time.time() - start_time) * 1000)

    # Update connector status based on test result
    connector.last_tested = datetime.utcnow()
    connector.last_sync_duration = duration_ms
    if success:
        connector.health_status = "Healthy"
        if connector.connector_type in ["Database", "LDAP", "API Gateway"]:
            connector.status = "Connected"
        connector.success_count = (connector.success_count or 0) + 1
    else:
        connector.health_status = "Unhealthy"
        if connector.connector_type in ["Database", "LDAP", "API Gateway"]:
            connector.status = "Failed"
        connector.failure_count = (connector.failure_count or 0) + 1

    db.commit()

    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Test Connection",
        details=message,
        status_val="Success" if success else "Failed"
    )

    return {
        "success": success,
        "message": message,
        "duration_ms": duration_ms,
        "tested_at": connector.last_tested.isoformat()
    }
@router.get("/connectors/{id}/tables")
def get_database_tables(id: int, db: Session = Depends(get_db)):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")
    if connector.connector_type != "Database":
        raise HTTPException(status_code=400, detail="Table listing is only available for Database connectors.")
    if connector.database_type != "MySQL":
        raise HTTPException(status_code=400, detail=f"Table listing for {connector.database_type} is not yet supported. Only MySQL is currently supported.")

    import pymysql
    decrypted_pw = decrypt_password(connector.password) if connector.password else ""
    try:
        conn = pymysql.connect(
            host=connector.host,
            port=connector.port or 3306,
            user=connector.username,
            password=decrypted_pw,
            database=connector.database_name,
            connect_timeout=connector.connection_timeout or 30
        )
        cursor = conn.cursor()
        cursor.execute("SHOW TABLES")
        tables = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return {"tables": tables}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to list tables: {str(e)}")


@router.get("/connectors/{id}/schema")
def get_connector_schema(
    id: int,
    table_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, connector)

    fields = []

    try:
        if connector.connector_type == "CSV":
            if not connector.file_path:
                raise Exception("No CSV file has been uploaded to this connector yet.")
            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                connector.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {connector.file_path}")

            import csv as csv_module
            with open(full_path, "r", encoding=connector.csv_encoding or "UTF-8") as f:
                reader = csv_module.reader(f, delimiter=connector.csv_delimiter or ",")
                header = next(reader, None)
                if not header:
                    raise Exception("CSV file appears to be empty or has no header row.")
                sample_row = next(reader, None)
                for idx, col_name in enumerate(header):
                    sample_val = sample_row[idx] if sample_row and idx < len(sample_row) else None
                    fields.append({
                        "field_name": col_name.strip(),
                        "data_type": "String",
                        "sample_value": sample_val
                    })

        elif connector.connector_type == "Excel":
            if not connector.file_path:
                raise Exception("No Excel file has been uploaded to this connector yet.")
            full_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                connector.file_path
            )
            if not os.path.exists(full_path):
                raise Exception(f"File not found on server: {connector.file_path}")

            wb = openpyxl.load_workbook(full_path, read_only=True)
            sheet = wb[connector.excel_sheet_name] if connector.excel_sheet_name and connector.excel_sheet_name in wb.sheetnames else wb.active

            rows_iter = sheet.iter_rows(max_row=2, values_only=True)
            header_row = next(rows_iter, None)
            sample_row = next(rows_iter, None)
            if not header_row:
                raise Exception("Excel sheet appears to be empty.")

            for idx, col_name in enumerate(header_row):
                if col_name is None:
                    continue
                sample_val = sample_row[idx] if sample_row and idx < len(sample_row) else None
                fields.append({
                    "field_name": str(col_name).strip(),
                    "data_type": "String",
                    "sample_value": str(sample_val) if sample_val is not None else None
                })

        elif connector.connector_type == "Database":
            if connector.database_type != "MySQL":
                raise Exception(f"Schema discovery for {connector.database_type} is not yet supported. Only MySQL is currently supported.")
            if not table_name:
                raise Exception("A table_name query parameter is required to discover schema for a Database connector.")

            # Save the selected database table_name in the unused file_path column
            # so that it can be used during scheduled imports.
            connector.file_path = table_name
            db.commit()

            import pymysql
            decrypted_pw = decrypt_password(connector.password) if connector.password else ""
            conn = pymysql.connect(
                host=connector.host,
                port=connector.port or 3306,
                user=connector.username,
                password=decrypted_pw,
                database=connector.database_name,
                connect_timeout=connector.connection_timeout or 30
            )
            cursor = conn.cursor()
            cursor.execute(f"DESCRIBE `{table_name}`")
            rows = cursor.fetchall()
            cursor.execute(f"SELECT * FROM `{table_name}` LIMIT 1")
            sample_row = cursor.fetchone()
            col_names = [desc[0] for desc in cursor.description]
            cursor.close()
            conn.close()

            for row in rows:
                col_name = row[0]
                col_type = row[1]
                sample_val = None
                if sample_row and col_name in col_names:
                    sample_val = sample_row[col_names.index(col_name)]
                fields.append({
                    "field_name": col_name,
                    "data_type": col_type,
                    "sample_value": str(sample_val) if sample_val is not None else None
                })

        elif connector.connector_type == "LDAP":
            raise Exception("Schema discovery for LDAP is not yet implemented.")

        elif connector.connector_type == "API Gateway":
            import requests as requests_lib
            
            url = connector.host
            if not url:
                raise Exception("API Gateway URL is required.")
            
            headers = {}
            if connector.file_path:
                try:
                    headers = json.loads(connector.file_path)
                except Exception:
                    pass

            auth = None
            if connector.auth_type == "Basic":
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                auth = (connector.username, decrypted_pw)
            elif connector.auth_type == "API Key" and connector.username and connector.password:
                decrypted_pw = decrypt_password(connector.password) if connector.password else ""
                headers[connector.username] = decrypted_pw

            timeout = connector.connection_timeout or 30
            resp = requests_lib.get(url, headers=headers, auth=auth, timeout=timeout)
            if resp.status_code >= 400:
                raise Exception(f"HTTP Error {resp.status_code}: {resp.text[:100]}")
            
            res_json = resp.json()
            json_key = connector.database_name
            target_list = res_json
            if json_key:
                if json_key in res_json:
                    target_list = res_json[json_key]
            
            if not isinstance(target_list, list) or len(target_list) == 0:
                # If it's a single dictionary, wrap it as a list to discover schema from it
                if isinstance(target_list, dict):
                    target_list = [target_list]
                else:
                    raise Exception(f"Expected a JSON array or object under JSON path key '{json_key or 'root'}'.")

            sample_record = target_list[0]
            if not isinstance(sample_record, dict):
                raise Exception("Discovered records are not JSON objects.")

            for key, val in sample_record.items():
                fields.append({
                    "field_name": str(key).strip(),
                    "data_type": type(val).__name__.capitalize() if val is not None else "String",
                    "sample_value": str(val) if val is not None else None
                })

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Schema Discovery",
        details=f"Discovered {len(fields)} field(s)" + (f" from table '{table_name}'" if table_name else ""),
        status_val="Success"
    )

    return {"fields": fields, "field_count": len(fields)}
@router.put("/connectors/{id}/schedule")
def update_connector_schedule(
    id: int,
    schedule_enabled: bool,
    schedule_frequency: str = None,
    schedule_time: str = None,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System")
):
    from app.services.scheduler import register_connector_schedule, unregister_connector_schedule, calculate_next_run

    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    if schedule_enabled and not schedule_frequency:
        raise HTTPException(status_code=400, detail="A schedule frequency is required when enabling scheduling.")
    if schedule_frequency and schedule_frequency not in ["Hourly", "Daily", "Weekly"]:
        raise HTTPException(status_code=400, detail="Frequency must be one of: Hourly, Daily, Weekly.")

    connector.schedule_enabled = schedule_enabled
    connector.schedule_frequency = schedule_frequency if schedule_enabled else None
    connector.schedule_time = schedule_time if (schedule_enabled and schedule_frequency in ["Daily", "Weekly"]) else None
    connector.modified_by = x_user_name
    connector.updated_at = datetime.utcnow()

    if schedule_enabled:
        register_connector_schedule(connector.id, schedule_frequency, schedule_time)
        connector.next_scheduled_run = calculate_next_run(schedule_frequency, schedule_time)
    else:
        unregister_connector_schedule(connector.id)
        connector.next_scheduled_run = None

    db.commit()
    db.refresh(connector)

    time_str = f" at {schedule_time}" if (schedule_time and schedule_frequency in ["Daily", "Weekly"]) else ""
    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Schedule Updated",
        details=f"Scheduling {'enabled (' + schedule_frequency + time_str + ')' if schedule_enabled else 'disabled'} by {x_user_name}.",
        status_val="Success"
    )

    return {
        "schedule_enabled": connector.schedule_enabled,
        "schedule_frequency": connector.schedule_frequency,
        "schedule_time": connector.schedule_time,
        "next_scheduled_run": connector.next_scheduled_run.isoformat() if connector.next_scheduled_run else None
    }


@router.post("/connectors/{id}/import")
def import_connector_data(
    id: int,
    table_name: Optional[str] = None,
    db: Session = Depends(get_db),
    x_user_name: str = Header(default="System"),
    x_user_role: str = Header(default="Read Only User")
):
    # RBAC check (Administrators and Data Stewards can run imports)
    if x_user_role not in ["Platform Administrator", "Data Steward"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. Only Platform Administrators and Data Stewards can import data."
        )

    # 1. Fetch connector
    connector = db.query(Connector).filter(Connector.id == id, Connector.is_deleted == False).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Connector not found")

    from app.utils.file_storage import restore_file_from_db_if_needed
    restore_file_from_db_if_needed(db, connector)

    # 2. Check mappings
    from app.models.connector_field_mapping import ConnectorFieldMapping
    mappings = db.query(ConnectorFieldMapping).filter(ConnectorFieldMapping.connector_id == id).all()
    if not mappings:
        raise HTTPException(
            status_code=400,
            detail="No attribute mappings found for this connector. Configure mappings first."
        )

    start_time = time.time()
    
    # 3. Read raw records (with a high limit, e.g. 2000, or None to read all)
    from app.services.preview_engine import PreviewEngine
    try:
        raw_rows = PreviewEngine._read_raw_records(connector, table_name, limit=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read source data: {str(e)}")

    if not raw_rows:
        return {"message": "No records found to import", "processed": 0, "imported": 0, "errors": 0}

    # 4. Load transform and validation engines and tables
    from app.models.transformation_rule import TransformationRule
    from app.models.validation_rule import ValidationRule
    from app.services.transformation_engine import TransformationEngine
    from app.services.validation_engine import ValidationEngine

    transformations = db.query(TransformationRule).filter(
        TransformationRule.connector_id == id,
        TransformationRule.enabled == True,
        TransformationRule.is_deleted == False
    ).order_by(TransformationRule.execution_order.asc()).all()

    validations = db.query(ValidationRule).filter(
        ValidationRule.connector_id == id,
        ValidationRule.enabled == True,
        ValidationRule.is_deleted == False
    ).order_by(ValidationRule.execution_order.asc()).all()

    mapping_by_id = {m.id: m for m in mappings}
    source_to_target = {m.source_field: m.target_attribute_name for m in mappings}
    
    seen_values_by_rule = {rule.id: set() for rule in validations if rule.validation_type.lower() == "unique"}

    # Process and filter out records that have validation errors
    parsed_identities = []
    
    total_processed = 0
    valid_count = 0
    warning_count = 0
    error_count = 0

    for index, raw_row in enumerate(raw_rows):
        total_processed += 1
        record_number = index + 1

        # Map raw fields to initial target attributes
        original_mapped = {}
        for source_f, val in raw_row.items():
            target_attr = source_to_target.get(source_f)
            if target_attr:
                original_mapped[target_attr] = str(val) if val is not None else ""

        # Fill in missing mapped fields as empty
        for m in mappings:
            if m.target_attribute_name not in original_mapped:
                original_mapped[m.target_attribute_name] = ""

        transformed = dict(original_mapped)

        # Apply transformations
        for trans_rule in transformations:
            if trans_rule.mapping_id in mapping_by_id:
                mapping = mapping_by_id[trans_rule.mapping_id]
                target_attr = mapping.target_attribute_name
                current_val = transformed.get(target_attr, "")
                
                transformed_val = TransformationEngine.transform_value(
                    value=current_val,
                    rule_type=trans_rule.transformation_type,
                    expression=trans_rule.expression,
                    parameters_str=trans_rule.parameters,
                    row_data=raw_row
                )
                transformed[target_attr] = transformed_val

        # Apply validations
        record_has_errors = False
        record_has_warnings = False

        for val_rule in validations:
            if val_rule.mapping_id in mapping_by_id:
                mapping = mapping_by_id[val_rule.mapping_id]
                target_attr = mapping.target_attribute_name
                transformed_val = transformed.get(target_attr, "")

                seen_set = seen_values_by_rule.get(val_rule.id)
                val_res = ValidationEngine.validate_value(
                    value=transformed_val,
                    validation_type=val_rule.validation_type,
                    parameters_str=val_rule.parameters,
                    error_message=val_rule.error_message,
                    seen_values=seen_set,
                    severity=val_rule.severity
                )

                if not val_res["valid"]:
                    status_sev = val_res["status"]
                    if status_sev == "Error":
                        record_has_errors = True
                    elif status_sev == "Warning":
                        record_has_warnings = True

        if record_has_errors:
            error_count += 1
            continue  # Skip inserting this record if there are validation errors
        
        if record_has_warnings:
            warning_count += 1
        
        valid_count += 1

        # Check if mapped to "Identity" module and build a real Identity record
        is_identity_mapped = any(m.target_module == "Identity" for m in mappings)
        if is_identity_mapped:
            employee_id_val = transformed.get("employee_id") or None
            first_name_val = transformed.get("first_name") or None
            last_name_val = transformed.get("last_name") or None
            display_name_val = transformed.get("display_name") or None
            if not display_name_val and (first_name_val or last_name_val):
                display_name_val = f"{first_name_val or ''} {last_name_val or ''}".strip()

            # Note: unlike the old logic, we do NOT invent a fake fallback email
            # (e.g. "user@ranalyzer.io") — a made-up email would silently break
            # the Identity Repository's account correlation, which matches on
            # real email values against imported Application accounts.
            email_val = transformed.get("email") or None
            department_val = transformed.get("department") or None
            job_title_val = transformed.get("job_title") or None
            manager_val = transformed.get("manager") or None
            status_val = transformed.get("status") or "Active"

            parsed_identities.append({
                "employee_id": employee_id_val,
                "first_name": first_name_val,
                "last_name": last_name_val,
                "display_name": display_name_val,
                "email": email_val,
                "department": department_val,
                "job_title": job_title_val,
                "manager": manager_val,
                "status": status_val,
                "attributes": transformed
            })

    # 5. Save the valid identity records into the stable Identity Repository
    # table. Only records previously imported BY THIS CONNECTOR are replaced —
    # other connectors' imported identities are left untouched. This table is
    # completely separate from the old IdentityRecord/dashboard demo data, so
    # nothing here can be wiped by the Dashboard's random "Sync API" feature.
    from app.models.identity import Identity
    if parsed_identities:
        # Load all active identities in a single query to avoid N+1 query overhead over remote DB
        existing_identities = db.query(Identity).filter(Identity.is_deleted == False).all()
        identity_by_emp_id = {i.employee_id: i for i in existing_identities if i.employee_id}
        identity_by_email = {i.email.lower(): i for i in existing_identities if i.email}

        for item in parsed_identities:
            existing_identity = None
            if item["employee_id"] and item["employee_id"] in identity_by_emp_id:
                existing_identity = identity_by_emp_id[item["employee_id"]]
            elif item["email"] and item["email"].lower() in identity_by_email:
                existing_identity = identity_by_email[item["email"].lower()]

            if existing_identity:
                # Update existing identity
                existing_identity.first_name = item["first_name"]
                existing_identity.last_name = item["last_name"]
                existing_identity.display_name = item["display_name"]
                existing_identity.email = item["email"]
                existing_identity.department = item["department"]
                existing_identity.job_title = item["job_title"]
                existing_identity.manager = item["manager"]
                existing_identity.status = item["status"]
                existing_identity.attributes = item["attributes"]
                existing_identity.source_connector_id = id
                existing_identity.source_connector_name = connector.connector_name
                existing_identity.modified_by = x_user_name
                existing_identity.updated_at = datetime.utcnow()
            else:
                # Add new identity
                rec = Identity(
                    employee_id=item["employee_id"],
                    first_name=item["first_name"],
                    last_name=item["last_name"],
                    display_name=item["display_name"],
                    email=item["email"],
                    department=item["department"],
                    job_title=item["job_title"],
                    manager=item["manager"],
                    status=item["status"],
                    attributes=item["attributes"],
                    source_connector_id=id,
                    source_connector_name=connector.connector_name,
                    imported_at=datetime.utcnow(),
                    created_by=x_user_name,
                    modified_by=x_user_name
                )
                db.add(rec)
        db.commit()

    duration_ms = int((time.time() - start_time) * 1000)

    # 6. Update Connector statistics
    connector.last_sync = datetime.utcnow()
    connector.last_sync_duration = duration_ms
    if error_count > 0 and valid_count == 0:
        connector.health_status = "Unhealthy"
        connector.status = "Failed"
        connector.failure_count = (connector.failure_count or 0) + 1
    elif error_count > 0:
        connector.health_status = "Degraded"
        connector.status = "Connected"
        connector.success_count = (connector.success_count or 0) + 1
    else:
        connector.health_status = "Healthy"
        connector.status = "Connected"
        connector.success_count = (connector.success_count or 0) + 1
        
    db.commit()
    db.refresh(connector)

    # 7. Write Connector log entry
    status_val = "Success" if error_count == 0 else ("Failed" if valid_count == 0 else "Warning")
    log_details = f"Import execution finished in {duration_ms}ms. Total processed: {total_processed}, imported: {valid_count}, warnings: {warning_count}, rejected: {error_count}."
    write_connector_log(
        db=db,
        connector_id=connector.id,
        action="Import Run",
        details=log_details,
        status_val=status_val
    )

    # 8. Write Recent Activity feed
    activity = RecentActivity(
        user=x_user_name,
        action=f"Data Source Sync - {connector.connector_name} - {valid_count} identities imported ({error_count} errors)",
        status="success" if error_count == 0 else "warning",
        created_at=datetime.utcnow()
    )
    db.add(activity)
    db.commit()

    return {
        "success": True,
        "processed": total_processed,
        "imported": valid_count,
        "warnings": warning_count,
        "errors": error_count,
        "duration_ms": duration_ms,
        "last_sync": connector.last_sync.isoformat()
    }